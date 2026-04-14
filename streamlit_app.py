<FILE filename="streamlit_app (20).py" size="50123 bytes">
"""
Decathlon Product Lookup
Improvements (applied per user request):
 - Variation for "Other": now uses 'size' column from master (not 'variation'); shows '...' ONLY when size is empty/missing
 - Price_KES: always "100000" (no comma) in final template
 - Stock: always 0 in final template
 - Preview: shows full Primary Category path only; Additional Category completely removed from frontend preview
 - sizes.txt: loaded ONLY from project folder (bundled) — no upload required
 - Fashion size editing: full editable preview of ALL SKUs using st.data_editor + SelectboxColumn so you can fix wrong sizes instantly
 - Invalid sizes (not in sizes.txt): new "Size Status" column with ❌ — rows with missing sizes are clearly marked
 - Template download: NOW outputs ONLY the "Upload Template" sheet (all other sheets are stripped before saving)
 - Product name logic unchanged (still appends color with " - " when needed)
"""

import os, io, re, json, asyncio
from typing import Optional
import numpy as np
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

try:
  from groq import AsyncGroq, Groq as SyncGroq
  GROQ_AVAILABLE = True
except ImportError:
  GROQ_AVAILABLE = False

# 
st.set_page_config(page_title="Decathlon Product Lookup", page_icon="", layout="wide")
st.markdown("""
<style>
h1 { color: #0082C3; }
.tag {
  display:inline-block; background:#0082C3; color:white;
  border-radius:4px; padding:2px 8px; font-size:12px; margin:2px;
}
.ai-badge {
  display:inline-block; background:linear-gradient(90deg,#f55036,#ff8c00);
  color:white; border-radius:12px; padding:2px 10px;
  font-size:11px; font-weight:700; margin-left:6px;
}
.kw-badge {
  display:inline-block; background:#0082C3; color:white;
  border-radius:12px; padding:2px 10px;
  font-size:11px; font-weight:700; margin-left:6px;
}
</style>
""", unsafe_allow_html=True)

st.title("Decathlon Product Lookup")
st.markdown("Search by SKU number — view details, images, and **download a filled upload template**.")

# Constants 
IMAGE_COLS  = ["OG_image"] + [f"picture_{i}"for i in range(1, 11)]
TEMPLATE_PATH ="product-creation-template.xlsx"
DECA_CAT_PATH ="deca_cat.xlsx"
MASTER_PATH  ="Decathlon_Working_File_Split.csv"

MASTER_TO_TEMPLATE = {
  "product_name": "Name",
  "designed_for": "Description",
  "sku_num_sku_r3":"SellerSKU",
  # ParentSKU is derived dynamically (first SKU per model_code), not mapped directly
  "brand_name":  "Brand",
  "bar_code":   "GTIN_Barcode",
  "color":     "color",
  "model_label":  "model",
  "OG_image":   "MainImage",
  "picture_1":   "Image2",
  "picture_2":   "Image3",
  "picture_3":   "Image4",
  "picture_4":   "Image5",
  "picture_5":   "Image6",
  "picture_6":   "Image7",
  "picture_7":   "Image8",
}

SIZES_PATH = "sizes.txt"

# =============================================================================
# UK SIZE EXTRACTION
# =============================================================================

_UK_SIZE_PATTERNS = [
  re.compile(r'\bUK\s*(\d{1,2}(?:\.\d)?)\b', re.IGNORECASE),
  re.compile(r'\bUK\s*(\d{1,2}(?:\.\d)?)\s*[-–]\s*\d{1,2}', re.IGNORECASE),
]

_CHILDREN_AGE_PATTERN = re.compile(
  r'(\d{1,2})\s*-\s*(\d{1,2})\s*(?:years?|yrs?)',
  re.IGNORECASE,
)

def extract_uk_size(raw: str) -> Optional[str]:
  if not raw:
    return None
  cleaned = re.sub(r'"+', '', raw).strip()
  for pat in _UK_SIZE_PATTERNS:
    m = pat.search(cleaned)
    if m:
      return f"UK {m.group(1)}"
  return None


def parse_valid_sizes(path: str) -> list:
  """Load sizes.txt from project folder only (no upload needed)."""
  try:
    with open(path, "r", encoding="utf-8") as f:
      lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]
    return lines
  except FileNotFoundError:
    return []


CATEGORY_MATCH_FIELDS = [
  "family","type","department_label","nature_label",
  "proposed_brand_name","brand_name","color","channable_gender",
  "size","keywords","designed_for","business_weight","product_name",
]

GROQ_SYSTEM_CAT ="""You are a product categorization expert for a sports retailer.
Given a product description and candidate category paths, pick the {top_n} best matches.
Consider brand, product type, gender, sport, and age group.

Respond with JSON only:
{{
 "categories": [
  {{"category":"<full path>","score": 0.95}},
  ...
 ]
}}

Rules:
- Return exactly {top_n} categories ordered by confidence descending
- Only pick from the provided candidate list - never invent categories
- Scores are floats 0.0-1.0
- JSON only, nothing else"""

GROQ_SYSTEM_DESC ="""You are a product copywriter for a sports retail marketplace.
Given product details, write exactly 3 short bullet points (each max 15 words) that highlight
the key features a buyer cares about. Focus on: sport/use-case, key benefit or material, target user.
Do NOT start with"Our team"or"Our designers". Be specific — mention the product name or sport.
Respond with JSON only:
{{"bullets": ["bullet 1","bullet 2","bullet 3"]}}
JSON only, nothing else."""

# =============================================================================
# HELPERS
# =============================================================================

def _clean(val) -> str:
  if pd.isna(val) or str(val).strip() in ("","-","nan"):
    return""
  return str(val).strip()


def _format_gtin(val) -> str:
  raw = str(val).strip()
  if not raw or raw.lower() in ("nan",""):
    return""
  try:
    return str(int(float(raw)))
  except (ValueError, OverflowError):
    return raw


# =============================================================================
# DATA LOADING
# =============================================================================

@st.cache_data(show_spinner=False)
def load_reference_data(file_bytes: bytes):
  wb_bytes = io.BytesIO(file_bytes)
  df_cat = pd.read_excel(wb_bytes, sheet_name="category", dtype=str)
  df_cat.columns = [c.strip() for c in df_cat.columns]
  df_cat = df_cat[df_cat["export_category"].notna() & (df_cat["export_category"].str.strip() !="")]
  df_cat["export_category"]   = df_cat["export_category"].str.strip()
  df_cat["category_name_lower"] = df_cat["category_name"].str.lower().str.strip()
  df_cat["Category Path lower"] = df_cat["Category Path"].str.lower().fillna("")
  df_cat["_path_tokens"] = df_cat["Category Path lower"].apply(
    lambda p: set(re.findall(r"[a-z]+", p))
  )
  wb_bytes.seek(0)
  df_brands = pd.read_excel(wb_bytes, sheet_name="brands", dtype=str, header=0)
  df_brands.columns = ["brand_entry"]
  df_brands = df_brands[df_brands["brand_entry"].notna()].copy()
  df_brands["brand_entry"]   = df_brands["brand_entry"].str.strip()
  df_brands["brand_name_lower"] = (
    df_brands["brand_entry"].str.split("-", n=1).str[-1].str.lower().str.strip()
  )
  return df_cat, df_brands


@st.cache_data(show_spinner=False)
def load_master(file_bytes: bytes, is_csv: bool) -> pd.DataFrame:
  if is_csv:
    try:
      return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8")
    except UnicodeDecodeError:
      return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")
  return pd.read_excel(io.BytesIO(file_bytes), dtype=str)


# =============================================================================
# TF-IDF INDEX
# =============================================================================

def _path_to_doc(path: str) -> str:
  parts = path.split("/")
  return"".join(parts) +""+"".join(parts[-3:]) * 2


@st.cache_resource(show_spinner=False)
def build_tfidf_index(ref_bytes: bytes):
  df_cat, _ = load_reference_data(ref_bytes)
  all_paths = df_cat["Category Path"].dropna().astype(str).tolist()
  path_set  = set(all_paths)
  leaves   = [p for p in all_paths
         if not any(other.startswith(p +"/") for other in path_set)]
  docs    = [_path_to_doc(p) for p in leaves]
  vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
  matrix   = vectorizer.fit_transform(docs)
  path_to_export = dict(zip(df_cat["Category Path"], df_cat["export_category"]))
  return leaves, vectorizer, matrix, path_to_export


def tfidf_shortlist(queries: list, leaves, vectorizer, matrix, k: int = 30) -> list:
  qmat = vectorizer.transform(queries)
  sims = cosine_similarity(qmat, matrix)
  out = []
  for row in sims:
    top_idx = np.argsort(row)[::-1][:k]
    out.append([leaves[i] for i in top_idx if row[i] > 0])
  return out


# =============================================================================
# KEYWORD MATCHING — vectorised
# =============================================================================

def _build_query_string(row: pd.Series) -> str:
  parts = []
  for f in CATEGORY_MATCH_FIELDS:
    val = row.get(f,"")
    if pd.notna(val) and str(val).strip() not in ("","-","nan"):
      parts.append(str(val).strip().lower())
  return"".join(parts)


def keyword_match_batch(rows_df: pd.DataFrame, df_cat: pd.DataFrame) -> list:
  queries = [_build_query_string(row) for _, row in rows_df.iterrows()]

  cat_token_sets = df_cat["_path_tokens"].tolist()
  cat_depths   = df_cat["Category Path lower"].str.count("/").tolist()
  cat_names   = df_cat["category_name_lower"].tolist()
  cat_exports  = df_cat["export_category"].tolist()
  n_cats     = len(cat_exports)

  results = []
  for query in queries:
    if not query:
      results.append(("",""))
      continue
    q_tokens = set(re.findall(r"[a-z]+", query))
    scores = [
      len(q_tokens & cat_token_sets[j])
      + (2 if cat_names[j] in query else 0)
      + cat_depths[j] * 0.1
      for j in range(n_cats)
    ]
    top2 = sorted(range(n_cats), key=lambda j: scores[j], reverse=True)[:2]
    primary  = cat_exports[top2[0]] if scores[top2[0]] > 0 else""
    secondary = cat_exports[top2[1]] if len(top2) > 1 and scores[top2[1]] > 0 else""
    results.append((primary, secondary))
  return results


def keyword_match_category(row: pd.Series, df_cat: pd.DataFrame) -> tuple:
  return keyword_match_batch(pd.DataFrame([row]), df_cat)[0]


# =============================================================================
# VARIATION (UPDATED per request)
# =============================================================================

def get_variation(
  row: pd.Series,
  is_fashion: bool = True,
  valid_sizes: Optional[list] = None,
  size_override: Optional[str] = None,
) -> str:
  """
  Fashion: size column + UK extraction + sizes.txt validation
  Other:  size column (NEW) — '...' ONLY when size is empty
  """
  if not is_fashion:
    raw = re.sub(r'"+', '', str(row.get("size", ""))).strip().rstrip(".")
    if raw.lower() in ("", "nan", "no size", "none"):
      return "..."
    return raw

  # Fashion path (unchanged)
  raw = re.sub(r'"+', '', str(row.get("size", ""))).strip().rstrip(".")
  if raw.lower() in ("", "nan", "no size", "none"):
    return size_override or "..."

  if size_override:
    return size_override

  if valid_sizes:
    raw_upper = raw.upper()
    for s in valid_sizes:
      if s.upper() == raw_upper:
        return s

  uk = extract_uk_size(raw)
  if uk and valid_sizes:
    uk_upper = uk.upper()
    for s in valid_sizes:
      if s.upper() == uk_upper:
        return s
    return uk

  if valid_sizes:
    raw_lower = raw.lower()
    for s in valid_sizes:
      if s.lower() in raw_lower or raw_lower in s.lower():
        return s

  return raw


# =============================================================================
# SHORT DESCRIPTION (rule-based, instant)
# =============================================================================

GENDER_MAP = {
  "MEN'S": "Men", "WOMEN'S": "Women", "BOYS'": "Boys", "GIRLS'": "Girls",
  "MEN": "Men", "WOMEN": "Women", "UNISEX": "Unisex", "NO GENDER": "",
  "HORSE": "",
}

_QUALITY_KEYWORDS = [
  "comfortable", "comfort", "lightweight", "light weight", "durable", "durability",
  "breathable", "breathability", "flexible", "flexibility", "waterproof", "water-resistant",
  "quick-dry", "quick dry", "moisture-wicking", "wicking", "anti-odour", "anti-odor",
  "odour-resistant", "stretch", "stretchable", "supportive", "support", "cushioned",
  "cushioning", "padded", "ergonomic", "adjustable", "reflective", "insulated",
  "warm", "cool", "softness", "soft", "reinforced", "abrasion-resistant", "non-slip",
  "grip", "ventilated", "ventilation", "seamless", "compression", "packable",
  "ultra-light", "high-performance", "performance", "protection", "protective",
]

def _extract_quality_phrases(desc: str, max_phrases: int = 2) -> list:
  if not desc:
    return []
  found = []
  used_words: set = set()
  desc_lower = desc.lower()
  for kw in _QUALITY_KEYWORDS:
    if kw in desc_lower:
      idx = desc_lower.index(kw)
      snippet = desc[idx:]
      sentence_end = re.search(r'[.!?]', snippet)
      if sentence_end:
        snippet = snippet[:sentence_end.start()]
      words_after = snippet.split()[:6]
      phrase = " ".join(words_after).rstrip(".,;:- ")
      phrase_words = set(phrase.lower().split())
      if phrase and len(phrase_words) >= 3 and len(phrase_words & used_words) < 2:
        found.append(phrase.capitalize())
        used_words |= phrase_words
      if len(found) >= max_phrases:
        break
  return found


def rule_based_short_desc(row: pd.Series) -> str:
  bullets = []

  brand  = _clean(row.get("brand_name", "")).title()
  dept   = _clean(row.get("department_label", "")).replace("/", "·").title()
  ptype  = _clean(row.get("type", "")).title()
  sport  = dept if dept else ptype
  g_raw  = _clean(row.get("channable_gender", "")).split("|")[0].strip().upper()
  gender = GENDER_MAP.get(g_raw, g_raw.title())

  b1_parts = [p for p in [brand, sport, gender] if p]
  if b1_parts:
    bullets.append(" · ".join(b1_parts))

  desc = _clean(row.get("designed_for", ""))
  quality_phrases = _extract_quality_phrases(desc, max_phrases=2)
  if quality_phrases:
    bullets.append(" · ".join(quality_phrases))
  elif desc:
    sentences = [s.strip() for s in re.split(r"[.!?]", desc) if len(s.strip()) > 20]
    feature = next(
      (s for s in sentences if not re.match(r"our (team|design)", s, re.I)),
      sentences[0] if sentences else "",
    )
    if feature:
      trunc = feature[:110].rsplit(" ", 1)[0] if len(feature) > 110 else feature
      bullets.append(trunc.capitalize())

  color = _clean(row.get("color", "")).split("|")[0].strip().title()
  size  = re.sub(r'"+', "", _clean(row.get("size", ""))).strip().rstrip(".")
  nature = _clean(row.get("nature_label", "")).title()

  if color and size and size.lower() != "no size":
    bullets.append(f"Colour: {color} · Size: {size}")
  elif color and nature:
    bullets.append(f"{nature} · {color}")
  elif color:
    bullets.append(f"Colour: {color}")
  elif size and size.lower() != "no size":
    bullets.append(f"Size: {size}")
  elif nature:
    bullets.append(nature)

  if not bullets:
    return ""
  items = "".join(f"<li>{b}</li>" for b in bullets[:3])
  return f"<ul>{items}</ul>"


# =============================================================================
# AI MATCHING (TF-IDF -> Groq, all parallel)
# =============================================================================

async def _async_rerank(idx, query, candidates, client, model, top_n, sem, task_type="cat"):
  async with sem:
    try:
      if task_type =="cat":
        cand_list ="\n".join(f"- {c}"for c in candidates)
        sys_msg  = GROQ_SYSTEM_CAT.format(top_n=top_n)
        user_msg = f"Product: {query}\n\nCandidates:\n{cand_list}"
      else:
        sys_msg  = GROQ_SYSTEM_DESC
        user_msg = f"Product details: {query}"

      resp = await client.chat.completions.create(
        model=model,
        temperature=0.15,
        response_format={"type":"json_object"},
        messages=[
          {"role":"system","content": sys_msg},
          {"role":"user", "content": user_msg},
        ],
      )
      raw = resp.choices[0].message.content.strip()
      data = json.loads(raw)
      return idx, data
    except Exception as e:
      return idx, {"error": str(e)}


async def _parallel_tasks(items, client, model, sem, task_type):
  tasks = [
    _async_rerank(i, q, c, client, model, 2, sem, task_type)
    for i, (q, c) in enumerate(items)
  ]
  raw = await asyncio.gather(*tasks)
  return [r for _, r in sorted(raw, key=lambda x: x[0])]


def groq_batch(items, api_key, model, concurrency, task_type="cat"):
  async def _run():
    client = AsyncGroq(api_key=api_key)
    sem  = asyncio.Semaphore(concurrency)
    return await _parallel_tasks(items, client, model, sem, task_type)
  return asyncio.run(_run())


def ai_match_categories(rows_df, leaves, vectorizer, matrix, path_to_export,
            api_key, model, shortlist_k=30, concurrency=10):

  def _resolve(cat_path: str) -> str:
    if cat_path in path_to_export:
      return path_to_export[cat_path]
    for p, ex in path_to_export.items():
      if p.endswith(cat_path) or cat_path.endswith(p):
        return ex
    return cat_path

  model_to_query: dict = {}
  model_order: list  = []
  for _, row in rows_df.iterrows():
    mc = str(row.get("model_code","")).strip()
    if mc and mc not in model_to_query:
      group = rows_df[rows_df["model_code"] == mc]
      model_to_query[mc] = _build_query_string(group.iloc[0])
      model_order.append(mc)

  unique_queries = [model_to_query[mc] for mc in model_order]
  candidates_list = tfidf_shortlist(unique_queries, leaves, vectorizer, matrix, shortlist_k)
  items      = list(zip(unique_queries, candidates_list))
  raw_preds    = groq_batch(items, api_key, model, concurrency, task_type="cat")

  model_to_cats: dict = {}
  for mc, data in zip(model_order, raw_preds):
    cats   = data.get("categories", [])
    primary  = _resolve(cats[0]["category"]) if len(cats) > 0 else""
    secondary = _resolve(cats[1]["category"]) if len(cats) > 1 else""
    model_to_cats[mc] = (primary, secondary)

  results = []
  for _, row in rows_df.iterrows():
    mc = str(row.get("model_code","")).strip()
    if mc and mc in model_to_cats:
      results.append(model_to_cats[mc])
    else:
      q = _build_query_string(row)
      c = tfidf_shortlist([q], leaves, vectorizer, matrix, shortlist_k)[0]
      rd = groq_batch([(q, c)], api_key, model, 1, task_type="cat")[0]
      cats   = rd.get("categories", [])
      primary  = _resolve(cats[0]["category"]) if len(cats) > 0 else""
      secondary = _resolve(cats[1]["category"]) if len(cats) > 1 else""
      results.append((primary, secondary))

  return results, model_to_cats


def _build_desc_query_per_model(group_df: pd.DataFrame) -> str:
  row  = group_df.iloc[0]
  parts = [
    _clean(row.get("product_name","")),
    _clean(row.get("department_label","")),
    _clean(row.get("brand_name","")),
    _clean(row.get("channable_gender","")).split("|")[0].strip(),
    _clean(row.get("designed_for",""))[:300],
    _clean(row.get("keywords",""))[:100],
  ]
  return"|".join(p for p in parts if p)


def ai_short_descriptions(rows_df, api_key, model, concurrency=10):
  model_queries: dict = {}
  model_repr:  dict = {}
  for i, (_, row) in enumerate(rows_df.iterrows()):
    mc = str(row.get("model_code","")).strip()
    if mc and mc not in model_queries:
      group = rows_df[rows_df["model_code"] == mc]
      model_queries[mc] = _build_desc_query_per_model(group)
      model_repr[mc]  = i

  unique_models = list(model_queries.keys())
  items     = [(model_queries[mc], []) for mc in unique_models]
  raw_results  = groq_batch(items, api_key, model, concurrency, task_type="desc")

  model_to_desc: dict = {}
  for mc, data in zip(unique_models, raw_results):
    if "error"in data:
      fallback_row = rows_df.iloc[model_repr[mc]]
      model_to_desc[mc] = rule_based_short_desc(fallback_row)
    else:
      bullets = data.get("bullets", [])
      items ="".join(f"<li>{b}</li>"for b in bullets[:3])
      model_to_desc[mc] = f"<ul>{items}</ul>"

  descs = []
  for _, row in rows_df.iterrows():
    mc = str(row.get("model_code","")).strip()
    if mc and mc in model_to_desc:
      descs.append(model_to_desc[mc])
    else:
      descs.append(rule_based_short_desc(row))
  return descs


# =============================================================================
# BRAND MATCHING
# =============================================================================

def match_brand(raw: str, df_brands: pd.DataFrame) -> str:
  if not raw or pd.isna(raw):
    return""
  needle = str(raw).strip().lower()
  exact = df_brands[df_brands["brand_name_lower"] == needle]
  if not exact.empty:
    return exact.iloc[0]["brand_entry"]
  partial = df_brands[df_brands["brand_name_lower"].str.contains(needle, regex=False)]
  if not partial.empty:
    return partial.iloc[0]["brand_entry"]
  for _, brow in df_brands.iterrows():
    if brow["brand_name_lower"] in needle:
      return brow["brand_entry"]
  return str(raw).strip()


# =============================================================================
# TEMPLATE BUILDER (UPDATED — ONLY Upload Template sheet)
# =============================================================================

def build_template(
  results_df, df_cat, df_brands,
  ai_categories,
  short_descs,
  is_fashion: bool = True,
  valid_sizes: Optional[list] = None,
  size_override: Optional[str] = None,
  sku_to_size_override: Optional[dict] = None,
) -> bytes:
  wb = load_workbook(TEMPLATE_PATH)
  ws = wb["Upload Template"]

  header_map = {}
  for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    if val:
      header_map[val] = col_idx

  hfont   = ws.cell(row=1, column=1).font
  data_font = Font(name=hfont.name or"Calibri", size=hfont.size or 11)
  data_align = Alignment(vertical="center")

  model_to_first_sku: dict = {}
  for _, r in results_df.iterrows():
    mc = str(r.get("model_code","")).strip()
    sku = str(r.get("sku_num_sku_r3","")).strip()
    if mc and sku and mc not in model_to_first_sku:
      model_to_first_sku[mc] = sku

  if df_cat is not None:
    exp_to_fullpath: dict = {}
    for _, _cr in df_cat.iterrows():
      _exp = str(_cr.get("export_category", "")).strip()
      _fp  = str(_cr.get("Category Path", "")).strip()
      if _exp and _fp and _exp not in exp_to_fullpath:
        exp_to_fullpath[_exp] = _fp
  else:
    exp_to_fullpath = {}

  for i, (_, src_row) in enumerate(results_df.iterrows()):
    row_idx = i + 2
    row_data = {}

    for master_col, tmpl_col in MASTER_TO_TEMPLATE.items():
      val = src_row.get(master_col,"")
      if pd.notna(val) and str(val).strip() not in ("","nan"):
        row_data[tmpl_col] = str(val).strip()

    mc = str(src_row.get("model_code","")).strip()
    if mc and mc in model_to_first_sku:
      row_data["ParentSKU"] = model_to_first_sku[mc]

    gtin = _format_gtin(src_row.get("bar_code",""))
    if gtin:
      row_data["GTIN_Barcode"] = gtin

    product_name = str(src_row.get("product_name","")).strip()
    color_raw  = str(src_row.get("color","")).strip()
    color    = color_raw.split("|")[0].strip()

    if product_name and color:
      if not product_name.lower().endswith(color.lower()):
        row_data["Name"] = f"{product_name} - {color.title()}"
      else:
        row_data["Name"] = product_name
    elif product_name:
      row_data["Name"] = product_name

    bw = str(src_row.get("business_weight","")).strip()
    if bw and bw.lower() not in ("","nan"):
      row_data["product_weight"] = re.sub(r'\s*kg\s*$', '', bw, flags=re.IGNORECASE).strip()

    size_val = re.sub(r'"+', '', str(src_row.get("size",""))).strip().rstrip(".")
    if size_val.lower() not in ("","nan","no size"):
      pkg_name = row_data.get("Name", product_name)
      row_data["package_content"] = f"{pkg_name} - {size_val}"

    raw_brand = src_row.get("brand_name","")
    if pd.notna(raw_brand) and str(raw_brand).strip():
      row_data["Brand"] = match_brand(str(raw_brand), df_brands)

    if ai_categories and i < len(ai_categories):
      primary_code, secondary_code = ai_categories[i]
    else:
      primary_code, secondary_code = keyword_match_category(src_row, df_cat)

    primary_full  = exp_to_fullpath.get(primary_code, primary_code)
    secondary_full = exp_to_fullpath.get(secondary_code, secondary_code)

    if primary_full:
      row_data["PrimaryCategory"]   = primary_full
    if secondary_full:
      row_data["AdditionalCategory"] = secondary_full

    # Variation with per-SKU override support
    sku = str(src_row.get("sku_num_sku_r3", "")).strip()
    override_size = sku_to_size_override.get(sku) if sku_to_size_override else None
    row_data["variation"] = get_variation(
      src_row,
      is_fashion=is_fashion,
      valid_sizes=valid_sizes,
      size_override=override_size or size_override,
    )

    # NEW: always set these columns
    row_data["Price_KES"] = "100000"
    row_data["Stock"] = 0

    # Color family
    color_for_family = str(src_row.get("color", "")).strip()
    if color_for_family and color_for_family.lower() not in ("", "nan"):
      row_data["color_family"] = color_for_family.split("|")[0].strip()

    if short_descs and i < len(short_descs) and short_descs[i]:
      row_data["short_description"] = short_descs[i]

    for tmpl_col, value in row_data.items():
      if tmpl_col in header_map:
        cell      = ws.cell(row=row_idx, column=header_map[tmpl_col])
        cell.value   = value
        cell.font   = data_font
        cell.alignment = data_align

  # === CRITICAL CHANGE: Keep ONLY the "Upload Template" sheet ===
  for sheet_name in list(wb.sheetnames):
    if sheet_name != "Upload Template":
      wb.remove(wb[sheet_name])

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
  st.header("Master Data")
  uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx","csv"])

  st.markdown("---")
  st.header("Category Matching")
  use_ai_matching = st.toggle(
    "AI matching (Groq)",
    value=False,
    help="OFF = fast vectorised keyword/TF-IDF. ON = TF-IDF shortlist + Groq LLM rerank.",
  )

  if use_ai_matching:
    if not GROQ_AVAILABLE:
      st.error("Install groq: `pip install groq`")
      use_ai_matching = False
    else:
      st.markdown('<span class="ai-badge">AI MODE ON</span>', unsafe_allow_html=True)
      show_key   = st.checkbox("Show key while typing", value=False)
      groq_api_key = st.text_input(
        "Groq API key",
        type="default"if show_key else"password",
        value=os.environ.get("GROQ_API_KEY",""),
        placeholder="Paste your gsk_... key here",
      )
      if groq_api_key and not groq_api_key.startswith("gsk_"):
        st.warning("Groq keys usually start with `gsk_` — double-check.")
      st.caption("Free key at [console.groq.com](https://console.groq.com)")
      groq_model = st.selectbox(
        "Model",
        ["llama-3.1-8b-instant","llama-3.3-70b-versatile","mixtral-8x7b-32768"],
        index=0,
        help="8b = fastest & free. 70b = most accurate.",
      )
      shortlist_k = st.slider("Shortlist size (candidates/product)", 10, 50, 30)
      concurrency = st.slider("Parallel Groq requests", 1, 30, 10)
      st.markdown("---")
      ai_short_desc = st.toggle(
        "AI short descriptions (Groq)",
        value=True,
        help="Use Groq to generate 3 polished bullet points per product. OFF = instant rule-based bullets.",
      )
  else:
    st.markdown('<span class="kw-badge">KEYWORD MODE</span>', unsafe_allow_html=True)
    st.caption("Instant vectorised TF-IDF keyword matching. No API key needed.")
    groq_api_key =""
    groq_model  ="llama-3.1-8b-instant"
    shortlist_k  = 30
    concurrency  = 10
    ai_short_desc = False

  st.markdown("---")
  st.header("Product Type")
  product_type = st.radio(
    "Product type",
    ["Fashion", "Other"],
    index=0,
    horizontal=True,
    help=(
      "Fashion: uses the 'size' column with UK-size extraction + sizes.txt validation.\n"
      "Other: uses the 'size' column (NEW) — '...' only when size is empty."
    ),
  )
  is_fashion = product_type == "Fashion"

  # sizes.txt is bundled in project folder — NO upload needed
  valid_sizes: list = parse_valid_sizes(SIZES_PATH)
  if valid_sizes:
    st.sidebar.info(f"Bundled sizes.txt: {len(valid_sizes)} sizes")
  else:
    st.sidebar.warning("sizes.txt not found in project folder!")

  st.markdown("---")
  st.header("Search Fields")
  also_search_name = st.checkbox("Also search by product name", value=False)


# =============================================================================
# LOAD REFERENCE DATA
# =============================================================================

try:
  ref_bytes = open(DECA_CAT_PATH,"rb").read()
  st.sidebar.success("deca_cat.xlsx loaded")
except FileNotFoundError:
  ref_bytes = None
  st.sidebar.error(f"`{DECA_CAT_PATH}` not found. Place it alongside app.py and restart.")

if ref_bytes:
  df_cat, df_brands = load_reference_data(ref_bytes)
  st.sidebar.success(f"{len(df_cat):,} categories · {len(df_brands)} brands")
  leaves, vectorizer, tfidf_matrix, path_to_export = build_tfidf_index(ref_bytes)
else:
  df_cat = df_brands = leaves = vectorizer = tfidf_matrix = path_to_export = None


# =============================================================================
# LOAD MASTER DATA
# =============================================================================

master_bytes = None
is_csv    = True

if uploaded_master:
  master_bytes = uploaded_master.read()
  is_csv    = uploaded_master.name.endswith(".csv")
  df_master  = load_master(master_bytes, is_csv)
  st.sidebar.success(f"{len(df_master):,} product rows loaded")
else:
  loaded = False
  for path, csv in [(MASTER_PATH, True), (MASTER_PATH.replace(".csv",".xlsx"), False)]:
    try:
      master_bytes = open(path,"rb").read()
      is_csv    = csv
      df_master  = load_master(master_bytes, csv)
      st.sidebar.info(f"Bundled master · {len(df_master):,} rows")
      loaded = True
      break
    except FileNotFoundError:
      continue
  if not loaded:
    st.error("No master file found. Upload one in the sidebar.")
    st.stop()

img_cols_present = [c for c in IMAGE_COLS if c in df_master.columns]
data_cols    = [c for c in df_master.columns if c not in img_cols_present]


# =============================================================================
# SEARCH
# =============================================================================

def search(q: str) -> pd.DataFrame:
  mask = pd.Series(False, index=df_master.index)
  mask |= df_master["sku_num_sku_r3"].fillna("").str.strip() == q.strip()
  if also_search_name and"product_name"in df_master.columns:
    mask |= df_master["product_name"].fillna("").str.lower().str.contains(q.lower(), regex=False)
  return df_master[mask].copy()


# =============================================================================
# INPUT TABS
# =============================================================================

tab1, tab2 = st.tabs(["Upload a List", "Manual Entry"])
queries = []

with tab1:
  uploaded_list = st.file_uploader(
    "Upload file with SKU numbers",
    type=["xlsx","csv","txt"],
    help="One SKU per row. For Excel/CSV, SKUs must be in column A.",
  )
  if uploaded_list:
    ext = uploaded_list.name.rsplit(".", 1)[-1].lower()
    if ext =="txt":
      queries = [l.strip() for l in uploaded_list.read().decode().splitlines() if l.strip()]
    elif ext =="csv":
      q_df  = pd.read_csv(uploaded_list, header=None, dtype=str)
      queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
    else:
      q_df  = pd.read_excel(uploaded_list, header=None, dtype=str)
      queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
    st.success(f"Loaded **{len(queries)}** search terms")

with tab2:
  manual = st.text_area(
    "Enter one SKU number per line",
    height=160,
    placeholder="4273417\n4273418\n4273423",
  )
  if manual.strip():
    queries = [q.strip() for q in manual.strip().splitlines() if q.strip()]


# =============================================================================
# RESULTS
# =============================================================================

if queries:
  st.markdown("---")
  all_result_frames = []
  no_match     = []

  for q in queries:
    res = search(q)
    if res.empty:
      no_match.append(q)
    else:
      res.insert(0,"Search Term", q)
      all_result_frames.append((q, res))

  if no_match:
    st.warning(f"No matches found for: **{', '.join(no_match)}**")

  if all_result_frames:
    total_rows = sum(len(r) for _, r in all_result_frames)
    st.success(f"**{total_rows} rows** matched across **{len(all_result_frames)}** query(ies)")

    combined = pd.concat([r for _, r in all_result_frames], ignore_index=True)

    # 1. Category matching 
    ai_categories = None

    if df_cat is not None and use_ai_matching and groq_api_key:
      n        = len(combined)
      unique_models_n = combined["model_code"].nunique() if"model_code"in combined.columns else n
      est       = max(2, unique_models_n // concurrency + 2)
      with st.spinner(f"AI category matching {unique_models_n} unique models (~{est}s)…"):
        try:
          ai_categories, _model_cats = ai_match_categories(
            combined, leaves, vectorizer, tfidf_matrix, path_to_export,
            groq_api_key, groq_model, shortlist_k, concurrency,
          )
          st.success(f"AI matched {unique_models_n} models → {n} SKUs")
        except Exception as e:
          st.error(f"Groq category error: {e}")
          use_ai_matching = False
    elif df_cat is not None and use_ai_matching and not groq_api_key:
      st.warning("Enter your Groq API key in the sidebar to use AI matching.")
      use_ai_matching = False

    # 2. Short descriptions 
    short_descs = None

    if use_ai_matching and ai_short_desc and groq_api_key:
      with st.spinner(f"Generating AI short descriptions ({len(combined)} products)…"):
        try:
          short_descs = ai_short_descriptions(combined, groq_api_key, groq_model, concurrency)
          st.success("Short descriptions generated")
        except Exception as e:
          st.error(f"Short desc error: {e}")
          short_descs = None

    if short_descs is None:
      short_descs = [rule_based_short_desc(row) for _, row in combined.iterrows()]

    # 3. Results full data view + EDITABLE SIZES (Fashion only)
    st.markdown("---")
    st.subheader(f"Results — {total_rows} SKU(s) — Preview & Edit Sizes")

    preview = combined.copy()
    preview["_variation"] = preview.apply(
      lambda r: get_variation(r, is_fashion=is_fashion, valid_sizes=valid_sizes, size_override=None),
      axis=1,
    )
    preview["_short_description"] = short_descs if short_descs else [
      rule_based_short_desc(r) for _, r in preview.iterrows()
    ]

    if df_cat is not None:
      _exp_to_path = {}
      for _, _rc in df_cat.iterrows():
        _e = str(_rc.get("export_category","")).strip()
        _p = str(_rc.get("Category Path","")).strip()
        if _e and _p and _e not in _exp_to_path:
          _exp_to_path[_e] = _p
    else:
      _exp_to_path = {}

    def _code_to_path(code):
      return _exp_to_path.get(str(code).strip(), code) if code else""

    if ai_categories:
      preview["_primary_cat"]  = [_code_to_path(c[0]) for c in ai_categories]
      preview["_secondary_cat"] = [_code_to_path(c[1]) for c in ai_categories]
    elif df_cat is not None:
      kw = keyword_match_batch(preview, df_cat)
      preview["_primary_cat"]  = [_code_to_path(c[0]) for c in kw]
      preview["_secondary_cat"] = [_code_to_path(c[1]) for c in kw]
    else:
      preview["_primary_cat"]  =""
      preview["_secondary_cat"] =""

    # Primary Category only — Additional Category removed from preview
    priority_cols = ["sku_num_sku_r3","product_name","color","size",
            "brand_name","department_label","bar_code",
            "_variation","_primary_cat","_short_description"]

    if is_fashion and valid_sizes:
      preview["_size_status"] = preview["_variation"].apply(
        lambda v: "✅ Valid" if str(v) in valid_sizes or str(v) == "..." else f"❌ Missing in sizes.txt"
      )
      priority_cols.insert(priority_cols.index("_variation") + 1, "_size_status")

    extra_cols = [c for c in data_cols if c not in priority_cols and c !="Search Term"]
    show_cols = [c for c in priority_cols if c in preview.columns] + [
      c for c in extra_cols if c in preview.columns
    ]

    variation_label = "Size (validated)" if is_fashion else "Variation (from size)"

    # Column config (shared)
    column_config_dict = {
      "sku_num_sku_r3":    st.column_config.TextColumn("SKU",        width="small"),
      "product_name":     st.column_config.TextColumn("Product",       width="large"),
      "color":        st.column_config.TextColumn("Colour",       width="medium"),
      "size":         st.column_config.TextColumn("Size (master)",   width="medium"),
      "brand_name":      st.column_config.TextColumn("Brand",        width="small"),
      "department_label":   st.column_config.TextColumn("Department",     width="medium"),
      "bar_code":       st.column_config.TextColumn("Barcode",       width="medium"),
      "_variation":      st.column_config.TextColumn(variation_label,   width="medium"),
      "_primary_cat":     st.column_config.TextColumn("Primary Cat",     width="large"),
      "_short_description":  st.column_config.TextColumn("Short Desc",     width="large"),
      "designed_for":     st.column_config.TextColumn("Description",     width="large"),
      "keywords":       st.column_config.TextColumn("Keywords",       width="large"),
    }

    # Editable preview for Fashion (per request)
    sku_to_size_override = None
    if is_fashion and valid_sizes:
      column_config_dict["_variation"] = st.column_config.SelectboxColumn(
        variation_label,
        options=["..."] + valid_sizes,
        width="medium",
      )
      column_config_dict["_size_status"] = st.column_config.TextColumn("Size Status", width="small")

      for c in [col for col in show_cols if col not in column_config_dict]:
        column_config_dict[c] = st.column_config.TextColumn(c.replace("_"," ").title(), width="medium")

      edited_df = st.data_editor(
        preview[show_cols],
        use_container_width=True,
        hide_index=True,
        height=420,
        column_config=column_config_dict,
      )
      sku_to_size_override = {
        str(k).strip(): v 
        for k, v in zip(edited_df["sku_num_sku_r3"], edited_df["_variation"])
      }
      st.caption("✅ Sizes are now editable above. Changes will be used in the final template download.")
    else:
      st.dataframe(
        preview[show_cols],
        use_container_width=True,
        hide_index=True,
        height=420,
        column_config=column_config_dict,
      )

    # 4. Category editor (kept because useful; Additional still available here)
    if df_cat is not None:
      st.markdown("---")
      mode_label ="AI"if (use_ai_matching and ai_categories) else"Keyword"
      st.subheader(f"Category Editor — {mode_label}")
      st.caption(
        "Categories are shared across all SKUs with the same model code."
        "Edit one row per model — siblings update automatically on export."
      )

      export_to_path: dict = {}
      for _, row_c in df_cat.iterrows():
        exp = str(row_c.get("export_category","")).strip()
        path = str(row_c.get("Category Path","")).strip()
        if exp and path and exp not in export_to_path:
          export_to_path[exp] = path
      path_label_to_export: dict = {v: k for k, v in export_to_path.items()}

      def export_to_label(code: str) -> str:
        if not code:
          return""
        return export_to_path.get(code, code)

      def label_to_export(label: str) -> str:
        if not label or label =="(auto)":
          return""
        return path_label_to_export.get(label, label)

      all_path_labels   = sorted(export_to_path.values())
      all_labels_w_blank = ["(auto)"] + all_path_labels

      if"cat_overrides"not in st.session_state:
        st.session_state.cat_overrides = {}

      cat_search = st.text_input(
        "Filter category list",
        placeholder="e.g. football, running, kids...",
        key="cat_search",
      )
      q = cat_search.strip().lower()
      filtered_labels = (
        ["(auto)"] + [lbl for lbl in all_path_labels if q in lbl.lower()]
        if q else all_labels_w_blank
      )
      st.caption(
        f"{len(filtered_labels)-1} categories shown"
        + (f" matching '{cat_search}'"if q else"(all)")
        + f" · {len(st.session_state.cat_overrides)} model override(s)"
      )
      st.markdown("---")

      seen_models: set = set()
      hc1, hc2, hc3, hc4 = st.columns([2, 3, 3, 1])
      hc1.markdown("**Model · SKUs**")
      hc2.markdown("**Primary Category**")
      hc3.markdown("**Additional Category**")
      hc4.markdown("**Method**")

      for i, (_, prow) in enumerate(combined.iterrows()):
        mc = str(prow.get("model_code","")).strip()
        if mc in seen_models:
          continue
        seen_models.add(mc)

        sku_count = len(combined[combined["model_code"] == mc])
        name   = str(prow.get("product_name",""))[:40]

        if ai_categories:
          first_idx = next(
            j for j, (_, r) in enumerate(combined.iterrows())
            if str(r.get("model_code","")).strip() == mc
          )
          auto_prim_code, auto_addl_code = ai_categories[first_idx]
        else:
          auto_prim_code, auto_addl_code = keyword_match_category(prow, df_cat)

        auto_prim_label = export_to_label(auto_prim_code)
        auto_addl_label = export_to_label(auto_addl_code)

        override = st.session_state.cat_overrides.get(mc, {})
        cur_prim_label = export_to_label(override.get("primary", auto_prim_code)) or auto_prim_label
        cur_addl_label = export_to_label(override.get("additional", auto_addl_code)) or auto_addl_label

        c1, c2, c3, c4 = st.columns([2, 3, 3, 1])
        c1.markdown(f"**{mc}** \n{name} \n`{sku_count} SKU(s)`")

        prim_opts = (
          filtered_labels if cur_prim_label in filtered_labels
          else ["(auto)", cur_prim_label] + [l for l in filtered_labels if l !="(auto)"]
        )
        try:  prim_idx = prim_opts.index(cur_prim_label)
        except ValueError: prim_idx = 0
        new_prim_label = c2.selectbox(
          f"Primary #{mc}", prim_opts,
          index=prim_idx, label_visibility="collapsed", key=f"prim_{mc}",
        )

        addl_opts = (
          filtered_labels if cur_addl_label in filtered_labels
          else ["(auto)", cur_addl_label] + [l for l in filtered_labels if l !="(auto)"]
        )
        try:  addl_idx = addl_opts.index(cur_addl_label)
        except ValueError: addl_idx = 0
        new_addl_label = c3.selectbox(
          f"Additional #{mc}", addl_opts,
          index=addl_idx, label_visibility="collapsed", key=f"addl_{mc}",
        )

        new_prim_code = label_to_export(new_prim_label) if new_prim_label !="(auto)"else auto_prim_code
        new_addl_code = label_to_export(new_addl_label) if new_addl_label !="(auto)"else auto_addl_code

        if new_prim_label !="(auto)"or new_addl_label !="(auto)":
          st.session_state.cat_overrides[mc] = {
            "primary":  new_prim_code,
            "additional": new_addl_code,
          }
        elif mc in st.session_state.cat_overrides:
          del st.session_state.cat_overrides[mc]

        badge ="Manual"if mc in st.session_state.cat_overrides else (
          "AI"if (use_ai_matching and ai_categories) else"Keyword"
        )
        c4.markdown(f"`{badge}`")

    # 5. Download buttons 
    st.markdown("---")
    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
      raw_out = io.BytesIO()
      with pd.ExcelWriter(raw_out, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Results")
      st.download_button(
        "Download Raw Results (.xlsx)",
        data=raw_out.getvalue(),
        file_name="decathlon_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
      )

    with col_dl2:
      if df_cat is None:
        st.warning("deca_cat.xlsx not loaded — template download unavailable.")
      else:
        try:
          merged_cats = []
          for _, prow in combined.iterrows():
            mc    = str(prow.get("model_code","")).strip()
            override = st.session_state.get("cat_overrides", {}).get(mc)
            if override:
              merged_cats.append((override["primary"], override["additional"]))
            elif ai_categories:
              first_idx = next(
                j for j, (_, r) in enumerate(combined.iterrows())
                if str(r.get("model_code","")).strip() == mc
              )
              merged_cats.append(ai_categories[first_idx])
            else:
              merged_cats.append(keyword_match_category(prow, df_cat))

          tpl_bytes = build_template(
            combined, df_cat, df_brands,
            ai_categories=merged_cats,
            short_descs=short_descs,
            is_fashion=is_fashion,
            valid_sizes=valid_sizes,
            size_override=None,
            sku_to_size_override=sku_to_size_override,
          )
          st.download_button(
            "✅ Download Upload Template Sheet ONLY (.xlsx)",
            data=tpl_bytes,
            file_name="decathlon_upload_template_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
          )
          st.caption("The downloaded file contains **only** the Upload Template sheet (all other sheets removed).")
        except FileNotFoundError:
          st.warning(
            "Template file not found."
            "Place `product-creation-template.xlsx` in the app folder."
          )
else:
  st.info("Upload a list or type search terms above to get started.")

st.markdown("---")
st.caption("Decathlon Product Lookup · Powered by your Decathlon working file")
</FILE>
