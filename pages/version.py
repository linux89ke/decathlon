

"""
Decathlon Product Lookup

Fixes & improvements in this version:
 - Uses Jumia AI Gateway via direct HTTP (urllib), no openai package needed
 - Adds AI debug panel with status codes, raw gateway response, payload, and fallback info
 - extract_uk_size now captures full range (UK 20-22, not just UK 20)
 - Other mode: variation comes from 'size' column (not 'variation'), dots only when empty
 - Price_KES always 100,000; Stock always 0
 - Product name colour append checks if colour is anywhere in name
 - Primary Category shown in full on front-end; Additional Category hidden
 - sizes.txt loaded from project folder automatically
 - Front-end preview table shows final export look with per-row size override dropdowns
 - Rows with size missing from sizes.txt are highlighted red in preview
 - Template export writes only the Upload Template sheet
 - Images are packed sequentially, skipping blanks, non-HTTP, and duplicates
 - Pillow image validation: main image checked for dimensions; fallback used if invalid
 - Header mapping ignores spaces, underscores, and capitalization
 - Description uses build_long_description with fallback for short/missing text
 - Unused Size or Variation column is physically deleted from the template
 - Category exports as "CODE - FULL PATH" using numeric code only
 - Missing template columns are auto-created
 - AI JSON extraction is more robust
 - Clean color is explicitly exported to a dedicated 'Color' column
 - Gender normalization helpers added (normalize_channable_gender, gender_query_token, gender_desc_label)
 - _build_query_string uses _clean_category_text + gender_query_token
 - rule_based_short_desc uses gender_desc_label + build_long_description
"""

import os, io, re, json, asyncio
import urllib.request
import urllib.error
from typing import Optional

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image

AI_GATEWAY_AVAILABLE = True

st.set_page_config(page_title="Decathlon Product Lookup", page_icon="", layout="wide")

if "run_id" not in st.session_state:
    st.session_state["run_id"] = 1

st.markdown("""
<style>
h1 { color: #0082C3; }
.tag {
  display:inline-block; background:#0082C3; color:white;
  border-radius:4px; padding:2px 8px; font-size:12px; margin:2px;
}
.ai-badge {
  display:inline-block; background:linear-gradient(90deg,#f55036,#ff8c00);
  color:white; border-radius:12px; padding:2px 10px;
  font-size:11px; font-weight:700; margin-left:6px;
}
.kw-badge {
  display:inline-block; background:#0082C3; color:white;
  border-radius:12px; padding:2px 10px;
  font-size:11px; font-weight:700; margin-left:6px;
}
</style>
""", unsafe_allow_html=True)

st.title("Decathlon Product Lookup")
st.markdown("Search by SKU number, view details, images, and **download a filled upload template**.")

# =============================================================================
# CONSTANTS
# =============================================================================

IMAGE_COLS = [
    "OG_image",
    "picture_1",
    "image_1",
    "image_2",
    "picture_2",
    "picture_3",
    "picture_4",
    "picture_5",
    "picture_6",
    "picture_7",
    "picture_8",
    "picture_9",
    "picture_10",
]

TEMPLATE_PATH = "product-creation-template.xlsx"
DECA_CAT_PATH = "deca_cat.xlsx"
MASTER_PATH = "Decathlon Working File Split.xlsx"
SIZES_PATH = "sizes.txt"

FALLBACK_MAIN_IMAGE_URL = "https://vendorcenter.jumia.com/product-set-images/2025/05/06/gvc.product.image.1746540897966.0d6c9cdb-fd44-4b85-9f8e-0ba625592dd4.jpeg"
MIN_IMAGE_DIMENSION = 200
MAX_IMAGE_DIMENSION = 3000

AI_GATEWAY_BASE_URL = os.environ.get(
    "AI_GATEWAY_BASE_URL",
    "https://ai-gateway.zuma.jumia.com/v1"
)

# Paste your working key locally if you want it hardcoded temporarily.
HARDCODED_AI_GATEWAY_KEY = ""

DEFAULT_AI_MODEL = os.environ.get("AI_GATEWAY_MODEL", "gpt-4o-mini")
DEBUG_AI_DEFAULT = True

debug_ai = False
ai_cat_debug = []
ai_desc_debug = []

MASTER_TO_TEMPLATE = {
    "product_name": "Name",
    "sku_num_sku_r3": "SellerSKU",
    "brand_name": "Brand",
    "bar_code": "GTIN_Barcode",
    "model_label": "model",
}

TEMPLATE_IMAGE_SLOTS = ["MainImage"] + [f"Image{i}" for i in range(2, 9)]

CATEGORY_MATCH_FIELDS = [
    "family", "type", "department_label", "nature_label",
    "proposed_brand_name", "brand_name", "color", "channable_gender",
    "size", "keywords", "description", "business_weight", "product_name",
]

NOISY_TERMS = [
    "decathlon", "no size", "no gender",
]

AI_SYSTEM_CAT = """You are a product categorization expert for a sports retailer.
Given a product description and candidate category paths, pick the {top_n} best matches.
Consider brand, product type, gender, sport, and age group.

Respond with a valid JSON object EXACTLY like this:
{{
 "categories": [
  {{"category":"<full path>","score": 0.95}},
  ...
 ]
}}

Rules:
- Return exactly {top_n} categories ordered by confidence descending
- Only pick from the provided candidate list, never invent categories
- Scores are floats 0.0-1.0
- RETURN ONLY VALID JSON. No intro text, no markdown formatting, no explanations."""

AI_SYSTEM_DESC = """You are a product copywriter for a sports retail marketplace.
Given product details, write exactly 3 short bullet points (each max 15 words) that highlight
the key features a buyer cares about. Focus on sport or use-case, key benefit or material, and target user.
Do not start with "Our team" or "Our designers". Be specific, mention the product name or sport.

Respond with a valid JSON object EXACTLY like this:
{{"bullets": ["bullet 1","bullet 2","bullet 3"]}}

RETURN ONLY VALID JSON. No intro text, no markdown formatting, no explanations."""

# =============================================================================
# UK SIZE EXTRACTION
# =============================================================================

_UK_SIZE_PATTERNS = [
    re.compile(r'\b(UK\s*\d{1,2}(?:\.\d)?\s*[-\u2013]\s*\d{1,2})\b', re.IGNORECASE),
    re.compile(r'\b(UK\s*\d{1,2}(?:\.\d)?)\b', re.IGNORECASE),
]

_CHILDREN_AGE_PATTERN = re.compile(
    r'(\d{1,2})\s*-\s*(\d{1,2})\s*(?:years?|yrs?)',
    re.IGNORECASE,
)

def extract_uk_size(raw: str) -> Optional[str]:
    if not raw:
        return None
    cleaned = re.sub(r'"+', '', raw).strip()
    for pat in _UK_SIZE_PATTERNS:
        m = pat.search(cleaned)
        if m:
            val = re.sub(r'^(UK)\s*', 'UK ', m.group(1), flags=re.IGNORECASE)
            return val.strip()
    return None

def parse_valid_sizes(path: str) -> list:
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]
        return lines
    except FileNotFoundError:
        return []

# =============================================================================
# HELPERS
# =============================================================================

def _clean(val) -> str:
    if pd.isna(val) or str(val).strip() in ("", "-", "nan"):
        return ""
    return str(val).strip()

def _format_gtin(val) -> str:
    raw = str(val).strip()
    if not raw or raw.lower() in ("nan", ""):
        return ""
    try:
        return str(int(float(raw)))
    except (ValueError, OverflowError):
        return raw

def extract_export_code(export_category: str) -> str:
    if not export_category:
        return ""
    m = re.match(r'^\s*(\d+)', str(export_category))
    return m.group(1) if m else ""

def normalize_channable_gender(raw: str) -> str:
    if raw is None:
        return ""

    txt = str(raw).upper().strip()
    if txt in ("", "NAN", "NONE"):
        return ""

    txt = txt.replace("\u2019", "'")
    txt = re.sub(r'\s+', ' ', txt)

    parts = [p.strip() for p in re.split(r'[|,/]+', txt) if p.strip()]
    joined = " | ".join(parts) if parts else txt

    has_baby  = "BABY BOY" in joined or "BABY GIRL" in joined
    has_boy   = "BOYS'" in joined or "BOY" in joined
    has_girl  = "GIRLS'" in joined or "GIRL" in joined
    has_men   = "MEN" in joined or "MEN'S" in joined
    has_women = "WOMEN" in joined or "WOMEN'S" in joined
    has_unisex    = "UNISEX" in joined
    has_no_gender = "NO GENDER" in joined

    if has_baby:
        return "Baby"
    if has_boy and has_girl:
        return "Kids"
    if has_boy:
        return "Boys"
    if has_girl:
        return "Girls"
    if has_men and has_women:
        return "Unisex"
    if has_men:
        return "Men"
    if has_women:
        return "Women"
    if has_unisex:
        return "Unisex"
    if has_no_gender:
        return "No Gender"
    return ""

def gender_query_token(raw: str) -> str:
    g = normalize_channable_gender(raw)
    mapping = {
        "Men":      "men",
        "Women":    "women",
        "Boys":     "boys kids",
        "Girls":    "girls kids",
        "Kids":     "kids boys girls",
        "Baby":     "baby kids",
        "Unisex":   "unisex",
        "No Gender": "",
    }
    return mapping.get(g, "")

def gender_desc_label(raw: str) -> str:
    g = normalize_channable_gender(raw)
    return "" if g == "No Gender" else g

def build_long_description(row: pd.Series, min_len: int = 50) -> str:
    raw_desc = _clean(row.get("description", ""))
    if len(raw_desc) >= min_len:
        return raw_desc

    product_name = _clean(row.get("product_name", ""))
    brand  = _clean(row.get("brand_name", "")).title()
    dept   = _clean(row.get("department_label", "")).replace("/", " / ").title()
    ptype  = _clean(row.get("type", "")).title()
    gender = gender_desc_label(row.get("channable_gender", ""))

    parts = []

    if product_name:
        parts.append(product_name)

    details = [x for x in [brand, gender, dept or ptype] if x]
    if details:
        parts.append("Designed for " + ", ".join(details).lower())

    if raw_desc:
        parts.append(raw_desc)
    else:
        parts.append("A quality Decathlon product built for everyday use and sport performance.")

    text = ". ".join(p.strip(". ") for p in parts if p).strip()

    if len(text) < min_len:
        text += " Made for comfort, performance, and regular active use."

    return text.strip()

@st.cache_data(show_spinner=False)
def get_image_dimension_status(url: str) -> dict:
    if not url or not str(url).startswith(("http://", "https://")):
        return {"ok": False, "width": None, "height": None, "error": "invalid_url"}

    try:
        resp = requests.get(url, timeout=(10, 20))
        resp.raise_for_status()

        with Image.open(io.BytesIO(resp.content)) as img:
            width, height = img.size

        ok = (
            MIN_IMAGE_DIMENSION <= width  <= MAX_IMAGE_DIMENSION
            and MIN_IMAGE_DIMENSION <= height <= MAX_IMAGE_DIMENSION
        )

        return {"ok": ok, "width": width, "height": height, "error": ""}
    except Exception as e:
        return {"ok": False, "width": None, "height": None, "error": str(e)}

# =============================================================================
# DATA LOADING
# =============================================================================

@st.cache_data(show_spinner="Loading category and brand reference data...")
def load_reference_data(file_bytes: bytes):
    wb_bytes = io.BytesIO(file_bytes)
    df_cat = pd.read_excel(wb_bytes, sheet_name="category", dtype=str)
    df_cat.columns = [c.strip() for c in df_cat.columns]
    df_cat = df_cat[df_cat["export_category"].notna() & (df_cat["export_category"].str.strip() != "")]
    df_cat["export_category"] = df_cat["export_category"].str.strip()
    df_cat["category_name_lower"] = df_cat["category_name"].str.lower().str.strip()
    df_cat["Category Path lower"] = df_cat["Category Path"].str.lower().fillna("")
    df_cat["_path_tokens"] = df_cat["Category Path lower"].apply(
        lambda p: set(re.findall(r"[a-z]+", p))
    )

    wb_bytes.seek(0)
    df_brands = pd.read_excel(wb_bytes, sheet_name="brands", dtype=str, header=0)
    df_brands.columns = ["brand_entry"]
    df_brands = df_brands[df_brands["brand_entry"].notna()].copy()
    df_brands["brand_entry"] = df_brands["brand_entry"].str.strip()
    df_brands["brand_name_lower"] = (
        df_brands["brand_entry"].str.split("-", n=1).str[-1].str.lower().str.strip()
    )
    return df_cat, df_brands

_SKU_ALIASES = {"seller sku", "sellersku", "seller_sku", "sku", "sku_num_sku_r3"}

def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        if col.lower().replace(" ", "_") in _SKU_ALIASES and col != "sku_num_sku_r3":
            rename[col] = "sku_num_sku_r3"
            break
    if rename:
        df = df.rename(columns=rename)
    return df

MASTER_PICKLE_PATH = "master_data.pkl"

def _master_mtime() -> float:
    for path in [MASTER_PATH, MASTER_PATH.replace(".xlsx", ".csv")]:
        try:
            return os.path.getmtime(path)
        except OSError:
            continue
    return 0.0

@st.cache_data(show_spinner="Loading master product data...")
def load_master(file_bytes: bytes, is_csv: bool) -> pd.DataFrame:
    if is_csv:
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    return _normalise_columns(df)

def load_master_fast(is_uploaded: bool = False) -> pd.DataFrame:
    import pickle

    mtime = _master_mtime()

    if os.path.exists(MASTER_PICKLE_PATH):
        try:
            with open(MASTER_PICKLE_PATH, "rb") as f:
                cached = pickle.load(f)
            if cached.get("mtime") == mtime:
                return cached["df"]
        except Exception:
            pass

    df = None
    for path, csv in [(MASTER_PATH, False), (MASTER_PATH.replace(".xlsx", ".csv"), True)]:
        try:
            raw = open(path, "rb").read()
            df = load_master(raw, csv)
            break
        except FileNotFoundError:
            continue

    if df is None:
        return None

    try:
        with open(MASTER_PICKLE_PATH, "wb") as f:
            pickle.dump({"mtime": mtime, "df": df}, f)
    except Exception:
        pass

    return df

# =============================================================================
# TF-IDF INDEX
# =============================================================================

def _path_to_doc(path: str) -> str:
    parts = path.split("/")
    return " ".join(parts) + " " + " ".join(parts[-3:]) * 2

TFIDF_PICKLE_PATH = "tfidf_index.pkl"

def _cat_mtime() -> float:
    try:
        return os.path.getmtime(DECA_CAT_PATH)
    except OSError:
        return 0.0

@st.cache_resource(show_spinner="Loading category index...")
def build_tfidf_index(ref_bytes: bytes):
    import pickle
    from sklearn.feature_extraction.text import TfidfVectorizer

    mtime = _cat_mtime()

    if os.path.exists(TFIDF_PICKLE_PATH):
        try:
            with open(TFIDF_PICKLE_PATH, "rb") as f:
                cached = pickle.load(f)
            if cached.get("mtime") == mtime:
                return (
                    cached["leaves"],
                    cached["vectorizer"],
                    cached["matrix"],
                    cached["path_to_export"],
                )
        except Exception:
            pass

    df_cat, _ = load_reference_data(ref_bytes)
    all_paths = df_cat["Category Path"].dropna().astype(str).tolist()
    path_set  = set(all_paths)
    leaves    = [p for p in all_paths if not any(other.startswith(p + "/") for other in path_set)]
    docs      = [_path_to_doc(p) for p in leaves]
    vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
    matrix     = vectorizer.fit_transform(docs)
    path_to_export = dict(zip(df_cat["Category Path"], df_cat["export_category"]))

    try:
        with open(TFIDF_PICKLE_PATH, "wb") as f:
            pickle.dump({
                "mtime": mtime,
                "leaves": leaves,
                "vectorizer": vectorizer,
                "matrix": matrix,
                "path_to_export": path_to_export,
            }, f)
    except Exception:
        pass

    return leaves, vectorizer, matrix, path_to_export

def tfidf_shortlist(queries: list, leaves, vectorizer, matrix, k: int = 30) -> list:
    from sklearn.metrics.pairwise import cosine_similarity
    qmat = vectorizer.transform(queries)
    sims = cosine_similarity(qmat, matrix)
    out  = []
    for row in sims:
        top_idx = np.argsort(row)[::-1][:k]
        out.append([leaves[i] for i in top_idx if row[i] > 0])
    return out

# =============================================================================
# KEYWORD MATCHING
# =============================================================================

def _clean_category_text(text: str) -> str:
    if not text or str(text).strip().lower() in ("", "nan", "-"):
        return ""
    out = str(text).lower()
    for term in NOISY_TERMS:
        out = out.replace(term, " ")
    out = out.replace("|", " ")
    out = re.sub(r'[^a-z0-9\s/+\-]', ' ', out)
    out = re.sub(r'\s+', ' ', out).strip()
    return out

def _build_query_string(row: pd.Series) -> str:
    parts = [
        _clean_category_text(row.get("product_name", "")),
        _clean_category_text(row.get("type", "")),
        _clean_category_text(row.get("department_label", "")),
        _clean_category_text(row.get("nature_label", "")),
        _clean_category_text(row.get("brand_name", "")),
        gender_query_token(row.get("channable_gender", "")),
        _clean_category_text(row.get("description", ""))[:220],
        _clean_category_text(row.get("keywords", ""))[:120],
    ]
    return " ".join(p for p in parts if p)

def keyword_match_batch(rows_df: pd.DataFrame, df_cat: pd.DataFrame) -> list:
    queries       = [_build_query_string(row) for _, row in rows_df.iterrows()]
    cat_token_sets = df_cat["_path_tokens"].tolist()
    cat_depths    = np.array(df_cat["Category Path lower"].str.count("/").tolist(), dtype=np.float32)
    cat_names     = df_cat["category_name_lower"].tolist()
    cat_exports   = df_cat["export_category"].tolist()
    n_cats        = len(cat_exports)
    depth_bonus   = cat_depths * 0.1
    results       = []

    for query in queries:
        if not query:
            results.append(("", ""))
            continue
        q_tokens    = set(re.findall(r"[a-z]+", query))
        token_scores = np.array([len(q_tokens & s) for s in cat_token_sets], dtype=np.float32)
        name_bonus   = np.array([2.0 if n in query else 0.0 for n in cat_names], dtype=np.float32)
        scores       = token_scores + name_bonus + depth_bonus
        top2 = np.argpartition(scores, -min(2, n_cats))[-min(2, n_cats):]
        top2 = top2[np.argsort(scores[top2])[::-1]]
        primary   = cat_exports[top2[0]] if scores[top2[0]] > 0 else ""
        secondary = cat_exports[top2[1]] if len(top2) > 1 and scores[top2[1]] > 0 else ""
        results.append((primary, secondary))
    return results

def keyword_match_category(row: pd.Series, df_cat: pd.DataFrame) -> tuple:
    return keyword_match_batch(pd.DataFrame([row]), df_cat)[0]

# =============================================================================
# VARIATION
# =============================================================================

def get_variation(
    row: pd.Series,
    is_fashion: bool = True,
    valid_sizes: Optional[list] = None,
    size_override: Optional[str] = None,
) -> str:
    raw_size = re.sub(r'"+', '', str(row.get("size", ""))).strip().rstrip(".")

    if not is_fashion:
        if raw_size.lower() not in ("", "nan", "no size", "none"):
            return raw_size
        raw_var = re.sub(r'"+', '', str(row.get("variation", ""))).strip().rstrip(".")
        if raw_var.lower() not in ("", "nan", "no size", "none"):
            return raw_var
        return "..."

    if raw_size.lower() in ("", "nan", "no size", "none"):
        return size_override or "..."

    if size_override:
        return size_override

    if valid_sizes:
        raw_upper = raw_size.upper()
        for s in valid_sizes:
            if s.upper() == raw_upper:
                return s

    uk = extract_uk_size(raw_size)
    if uk and valid_sizes:
        uk_upper = uk.upper()
        for s in valid_sizes:
            if s.upper() == uk_upper:
                return s
        return uk

    if valid_sizes:
        raw_lower = raw_size.lower()
        for s in valid_sizes:
            if s.lower() in raw_lower or raw_lower in s.lower():
                return s

    return raw_size

@st.cache_data
def _valid_sizes_upper_set(sizes_tuple: tuple) -> frozenset:
    return frozenset(s.upper() for s in sizes_tuple)

def is_size_missing(computed_variation: str, valid_sizes: list) -> bool:
    if not valid_sizes:
        return False
    if computed_variation in ("...", ""):
        return True
    return computed_variation.upper() not in _valid_sizes_upper_set(tuple(valid_sizes))

# =============================================================================
# SHORT DESCRIPTION
# =============================================================================

GENDER_MAP = {
    "MEN'S": "Men", "WOMEN'S": "Women", "BOYS'": "Boys", "GIRLS'": "Girls",
    "MEN": "Men", "WOMEN": "Women", "UNISEX": "Unisex", "NO GENDER": "",
    "HORSE": "",
}

_QUALITY_KEYWORDS = [
    "comfortable", "comfort", "lightweight", "light weight", "durable", "durability",
    "breathable", "breathability", "flexible", "flexibility", "waterproof", "water-resistant",
    "quick-dry", "quick dry", "moisture-wicking", "wicking", "anti-odour", "anti-odor",
    "odour-resistant", "stretch", "stretchable", "supportive", "support", "cushioned",
    "cushioning", "padded", "ergonomic", "adjustable", "reflective", "insulated",
    "warm", "cool", "softness", "soft", "reinforced", "abrasion-resistant", "non-slip",
    "grip", "ventilated", "ventilation", "seamless", "compression", "packable",
    "ultra-light", "high-performance", "performance", "protection", "protective",
]

def _extract_quality_phrases(desc: str, max_phrases: int = 2) -> list:
    if not desc:
        return []
    found      = []
    used_words: set = set()
    desc_lower = desc.lower()
    for kw in _QUALITY_KEYWORDS:
        if kw in desc_lower:
            idx      = desc_lower.index(kw)
            snippet  = desc[idx:]
            sentence_end = re.search(r'[.!?]', snippet)
            if sentence_end:
                snippet = snippet[:sentence_end.start()]
            words_after = snippet.split()[:6]
            phrase      = " ".join(words_after).rstrip(".,;:- ")
            phrase_words = set(phrase.lower().split())
            if phrase and len(phrase_words) >= 3 and len(phrase_words & used_words) < 2:
                found.append(phrase.capitalize())
                used_words |= phrase_words
            if len(found) >= max_phrases:
                break
    return found

def rule_based_short_desc(row: pd.Series) -> str:
    bullets = []

    brand  = _clean(row.get("brand_name", "")).title()
    dept   = _clean(row.get("department_label", "")).replace("/", "·").title()
    ptype  = _clean(row.get("type", "")).title()
    sport  = dept if dept else ptype
    gender = gender_desc_label(row.get("channable_gender", ""))

    b1_parts = [p for p in [brand, gender, sport] if p]
    if b1_parts:
        bullets.append(" · ".join(b1_parts))

    desc = build_long_description(row)
    if desc:
        sentences = [s.strip() for s in re.split(r"[.!?]", desc) if len(s.strip()) > 20]
        feature   = sentences[0] if sentences else ""
        if feature:
            trunc = feature[:110].rsplit(" ", 1)[0] if len(feature) > 110 else feature
            bullets.append(trunc.capitalize())

    color  = _clean(row.get("color", "")).split("|")[0].strip().title()
    size   = re.sub(r'"+', "", _clean(row.get("size", ""))).strip().rstrip(".")
    nature = _clean(row.get("nature_label", "")).title()

    if color and size and size.lower() != "no size":
        bullets.append(f"Colour: {color} · Size: {size}")
    elif color and nature:
        bullets.append(f"{nature} · {color}")
    elif color:
        bullets.append(f"Colour: {color}")
    elif size and size.lower() != "no size":
        bullets.append(f"Size: {size}")
    elif nature:
        bullets.append(nature)

    if not bullets:
        return ""

    items_html = "".join(f"<li>{b}</li>" for b in bullets[:3])
    return f"<ul>{items_html}</ul>"

# =============================================================================
# AI MATCHING, DIRECT HTTP VERSION
# =============================================================================

def _safe_json_loads(text: str):
    try:
        return json.loads(text)
    except Exception:
        return None

def _parse_llm_json(raw: str):
    if not raw:
        return None

    parsed = _safe_json_loads(raw)
    if parsed is not None:
        return parsed

    md_ticks = '`' * 3
    clean = re.sub(r'^' + md_ticks + r'(?:json)?\s*', '', raw.strip(), flags=re.IGNORECASE)
    clean = re.sub(r'\s*' + md_ticks + r'$', '', clean.strip())
    parsed = _safe_json_loads(clean)
    if parsed is not None:
        return parsed

    match_obj = re.search(r'\{.*\}', raw, re.DOTALL)
    if match_obj:
        parsed = _safe_json_loads(match_obj.group(0))
        if parsed is not None:
            return parsed

    match_arr = re.search(r'\[.*\]', raw, re.DOTALL)
    if match_arr:
        parsed = _safe_json_loads(match_arr.group(0))
        if parsed is not None:
            return parsed

    return None

async def _gateway_chat_completion(api_key, model, messages, max_tokens=500, temperature=0.15, json_mode=True):
    def _post():
        payload = {
            "model": model,
            "messages": messages,
            "max_tokens": max_tokens,
        }

        if temperature is not None:
            payload["temperature"] = temperature

        if json_mode:
            payload["response_format"] = {"type": "json_object"}

        data = json.dumps(payload).encode("utf-8")
        url  = f"{AI_GATEWAY_BASE_URL.rstrip('/')}/chat/completions"

        req = urllib.request.Request(
            url=url,
            data=data,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            method="POST",
        )

        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                body   = resp.read().decode("utf-8", errors="replace")
                parsed = _safe_json_loads(body)
                return {
                    "ok": True,
                    "status": getattr(resp, "status", 200),
                    "url": url,
                    "body_text": body,
                    "body_json": parsed,
                    "request_payload": payload,
                }

        except urllib.error.HTTPError as e:
            body   = e.read().decode("utf-8", errors="replace")
            parsed = _safe_json_loads(body)
            return {
                "ok": False,
                "status": e.code,
                "url": url,
                "body_text": body,
                "body_json": parsed,
                "request_payload": payload,
                "error_type": "HTTPError",
                "error_message": str(e),
            }

        except urllib.error.URLError as e:
            return {
                "ok": False,
                "status": None,
                "url": url,
                "body_text": "",
                "body_json": None,
                "request_payload": payload,
                "error_type": "URLError",
                "error_message": str(e),
            }

        except Exception as e:
            return {
                "ok": False,
                "status": None,
                "url": url,
                "body_text": "",
                "body_json": None,
                "request_payload": payload,
                "error_type": type(e).__name__,
                "error_message": str(e),
            }

    return await asyncio.to_thread(_post)

def _extract_content_from_gateway_response(resp_dict):
    body_json = resp_dict.get("body_json")
    if not body_json:
        return None, "Gateway returned non-JSON response"

    try:
        content = body_json["choices"][0]["message"]["content"]
        if content is None:
            return None, "choices[0].message.content was null"
        return str(content).strip(), None
    except Exception as e:
        return None, f"Unexpected gateway response shape: {e}"

async def _async_rerank(idx, query, candidates, api_key, model, top_n, sem, task_type="cat", debug=False):
    async with sem:
        try:
            if task_type == "cat":
                cand_list = "\n".join(f"- {c}" for c in candidates)
                sys_msg   = AI_SYSTEM_CAT.format(top_n=top_n)
                user_msg  = f"Product: {query}\n\nCandidates:\n{cand_list}"
            else:
                sys_msg  = AI_SYSTEM_DESC
                user_msg = f"Product details: {query}"

            messages = [
                {"role": "system", "content": sys_msg},
                {"role": "user",   "content": user_msg},
            ]

            first_try = await _gateway_chat_completion(
                api_key=api_key, model=model, messages=messages,
                max_tokens=500, temperature=0.15, json_mode=True,
            )

            used_fallback = False
            final_resp    = first_try

            if not first_try.get("ok"):
                used_fallback = True
                final_resp = await _gateway_chat_completion(
                    api_key=api_key, model=model, messages=messages,
                    max_tokens=500, temperature=0.15, json_mode=False,
                )

            if not final_resp.get("ok"):
                return idx, {
                    "error": final_resp.get("error_message", "Gateway request failed"),
                    "debug": {
                        "task_type": task_type,
                        "used_fallback": used_fallback,
                        "first_try_status": first_try.get("status"),
                        "first_try_body": first_try.get("body_text", "")[:4000],
                        "status": final_resp.get("status"),
                        "url": final_resp.get("url"),
                        "error_type": final_resp.get("error_type"),
                        "error_message": final_resp.get("error_message"),
                        "body_text": final_resp.get("body_text", "")[:4000],
                        "request_payload": final_resp.get("request_payload"),
                        "query_preview": query[:400],
                        "candidate_count": len(candidates),
                        "candidate_preview": candidates[:10],
                        "model": model,
                    }
                }

            raw_content, extract_err = _extract_content_from_gateway_response(final_resp)
            if extract_err:
                return idx, {
                    "error": extract_err,
                    "debug": {
                        "task_type": task_type,
                        "used_fallback": used_fallback,
                        "status": final_resp.get("status"),
                        "url": final_resp.get("url"),
                        "body_text": final_resp.get("body_text", "")[:4000],
                        "request_payload": final_resp.get("request_payload"),
                        "query_preview": query[:400],
                        "candidate_count": len(candidates),
                        "candidate_preview": candidates[:10],
                        "model": model,
                    }
                }

            parsed = _parse_llm_json(raw_content)

            if parsed is None:
                return idx, {
                    "error": "Model content was not valid JSON",
                    "debug": {
                        "task_type": task_type,
                        "used_fallback": used_fallback,
                        "status": final_resp.get("status"),
                        "url": final_resp.get("url"),
                        "body_text": final_resp.get("body_text", "")[:4000],
                        "raw_model_content": raw_content[:4000],
                        "request_payload": final_resp.get("request_payload"),
                        "query_preview": query[:400],
                        "candidate_count": len(candidates),
                        "candidate_preview": candidates[:10],
                        "model": model,
                    }
                }

            if debug:
                parsed["_debug"] = {
                    "task_type": task_type,
                    "used_fallback": used_fallback,
                    "status": final_resp.get("status"),
                    "url": final_resp.get("url"),
                    "model": model,
                }

            return idx, parsed

        except Exception as e:
            return idx, {
                "error": str(e),
                "debug": {
                    "task_type": task_type,
                    "query_preview": query[:400],
                    "candidate_count": len(candidates),
                    "candidate_preview": candidates[:10],
                    "model": model,
                }
            }

async def _parallel_tasks(items, api_key, model, sem, task_type, debug=False):
    tasks = [
        _async_rerank(i, q, c, api_key, model, 2, sem, task_type, debug)
        for i, (q, c) in enumerate(items)
    ]
    raw = await asyncio.gather(*tasks)
    return [r for _, r in sorted(raw, key=lambda x: x[0])]

def gateway_batch(items, api_key, model, concurrency, task_type="cat", debug=False):
    async def _run():
        sem = asyncio.Semaphore(concurrency)
        return await _parallel_tasks(items, api_key, model, sem, task_type, debug)
    return asyncio.run(_run())

def ai_match_categories(rows_df, leaves, vectorizer, matrix, path_to_export,
                        api_key, model, shortlist_k=30, concurrency=10, debug=False):
    def _resolve(cat_path: str) -> str:
        if cat_path in path_to_export:
            return path_to_export[cat_path]
        for p, ex in path_to_export.items():
            if p.endswith(cat_path) or cat_path.endswith(p):
                return ex
        return cat_path

    model_to_query: dict = {}
    model_order:    list = []
    for _, row in rows_df.iterrows():
        mc = str(row.get("model_code", "")).strip()
        if mc and mc not in model_to_query:
            group = rows_df[rows_df["model_code"] == mc]
            model_to_query[mc] = _build_query_string(group.iloc[0])
            model_order.append(mc)

    unique_queries  = [model_to_query[mc] for mc in model_order]
    candidates_list = tfidf_shortlist(unique_queries, leaves, vectorizer, matrix, shortlist_k)
    items           = list(zip(unique_queries, candidates_list))
    raw_preds       = gateway_batch(items, api_key, model, concurrency, task_type="cat", debug=debug)

    model_to_cats: dict = {}
    debug_rows          = []

    for mc, data, query_used, cand_used in zip(model_order, raw_preds, unique_queries, candidates_list):
        if "error" in data:
            debug_rows.append({
                "model_code": mc,
                "task_type": "cat",
                "error": data.get("error", ""),
                "query_preview": query_used[:250],
                "candidate_count": len(cand_used),
                "debug": data.get("debug", {}),
            })
            model_to_cats[mc] = ("", "")
            continue

        cats = data if isinstance(data, list) else data.get("categories", []) if isinstance(data, dict) else []

        primary   = _resolve(cats[0].get("category", "")) if len(cats) > 0 and isinstance(cats[0], dict) else ""
        secondary = _resolve(cats[1].get("category", "")) if len(cats) > 1 and isinstance(cats[1], dict) else ""
        model_to_cats[mc] = (primary, secondary)

    results = []
    for _, row in rows_df.iterrows():
        mc = str(row.get("model_code", "")).strip()
        if mc and mc in model_to_cats:
            results.append(model_to_cats[mc])
        else:
            q  = _build_query_string(row)
            c  = tfidf_shortlist([q], leaves, vectorizer, matrix, shortlist_k)[0]
            rd = gateway_batch([(q, c)], api_key, model, 1, task_type="cat", debug=debug)[0]
            if "error" in rd:
                debug_rows.append({
                    "model_code": mc,
                    "task_type": "cat-single",
                    "error": rd.get("error", ""),
                    "query_preview": q[:250],
                    "candidate_count": len(c),
                    "debug": rd.get("debug", {}),
                })
                results.append(("", ""))
            else:
                cats      = rd if isinstance(rd, list) else rd.get("categories", []) if isinstance(rd, dict) else []
                primary   = _resolve(cats[0].get("category", "")) if len(cats) > 0 and isinstance(cats[0], dict) else ""
                secondary = _resolve(cats[1].get("category", "")) if len(cats) > 1 and isinstance(cats[1], dict) else ""
                results.append((primary, secondary))

    return results, model_to_cats, debug_rows

def _build_desc_query_per_model(group_df: pd.DataFrame) -> str:
    row   = group_df.iloc[0]
    parts = [
        _clean(row.get("product_name", "")),
        _clean(row.get("department_label", "")),
        _clean(row.get("brand_name", "")),
        _clean(row.get("channable_gender", "")).split("|")[0].strip(),
        _clean(row.get("description", ""))[:300],
        _clean(row.get("keywords", ""))[:100],
    ]
    return "|".join(p for p in parts if p)

def ai_short_descriptions(rows_df, api_key, model, concurrency=10, debug=False):
    model_queries: dict = {}
    model_repr:    dict = {}

    for i, (_, row) in enumerate(rows_df.iterrows()):
        mc = str(row.get("model_code", "")).strip()
        if mc and mc not in model_queries:
            group = rows_df[rows_df["model_code"] == mc]
            model_queries[mc] = _build_desc_query_per_model(group)
            model_repr[mc]    = i

    unique_models = list(model_queries.keys())
    items         = [(model_queries[mc], []) for mc in unique_models]
    raw_results   = gateway_batch(items, api_key, model, concurrency, task_type="desc", debug=debug)

    model_to_desc: dict = {}
    debug_rows          = []

    for mc, data in zip(unique_models, raw_results):
        if "error" in data:
            fallback_row      = rows_df.iloc[model_repr[mc]]
            model_to_desc[mc] = rule_based_short_desc(fallback_row)
            debug_rows.append({
                "model_code": mc,
                "task_type": "desc",
                "error": data.get("error", ""),
                "query_preview": model_queries[mc][:250],
                "candidate_count": 0,
                "debug": data.get("debug", {}),
            })
        else:
            bullets = (
                data if isinstance(data, list)
                else data.get("bullets", data.get("bullet_points", data.get("items", [])))
                if isinstance(data, dict) else []
            )
            if bullets and isinstance(bullets, list):
                items_ = "".join(f"<li>{b}</li>" for b in bullets[:3])
                model_to_desc[mc] = f"<ul>{items_}</ul>"
            else:
                fallback_row      = rows_df.iloc[model_repr[mc]]
                model_to_desc[mc] = rule_based_short_desc(fallback_row)

    descs = []
    for _, row in rows_df.iterrows():
        mc = str(row.get("model_code", "")).strip()
        descs.append(model_to_desc[mc] if mc and mc in model_to_desc else rule_based_short_desc(row))
    return descs, debug_rows

# =============================================================================
# BRAND MATCHING
# =============================================================================

def match_brand(raw: str, df_brands: pd.DataFrame) -> str:
    if not raw or pd.isna(raw):
        return ""
    needle = str(raw).strip().lower()
    exact  = df_brands[df_brands["brand_name_lower"] == needle]
    if not exact.empty:
        return exact.iloc[0]["brand_entry"]
    partial = df_brands[df_brands["brand_name_lower"].str.contains(needle, regex=False, na=False)]
    if not partial.empty:
        return partial.iloc[0]["brand_entry"]
    match = df_brands[df_brands["brand_name_lower"].apply(lambda b: b in needle)]
    if not match.empty:
        return match.iloc[0]["brand_entry"]
    return str(raw).strip()

# =============================================================================
# TEMPLATE BUILDER
# =============================================================================

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

def build_template(
    results_df,
    df_cat,
    df_brands,
    ai_categories,
    short_descs,
    is_fashion: bool = True,
    valid_sizes: Optional[list] = None,
    size_overrides: Optional[dict] = None,
) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)

    sheets_to_remove = [s for s in wb.sheetnames if s != "Upload Template"]
    for s in sheets_to_remove:
        del wb[s]

    ws = wb["Upload Template"]

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            norm_val = re.sub(r'[^a-z0-9]', '', str(val).lower())
            header_map[norm_val] = col_idx

    unused_col_key = "variation" if is_fashion else "size"
    if unused_col_key in header_map:
        ws.delete_cols(header_map[unused_col_key])

        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                norm_val = re.sub(r'[^a-z0-9]', '', str(val).lower())
                header_map[norm_val] = col_idx

    active_col_key   = "size" if is_fashion else "variation"
    active_col_label = "Size" if is_fashion else "Variation"

    if active_col_key not in header_map:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = active_col_label
        header_map[active_col_key] = new_col

    current_max_col = ws.max_column

    hfont      = ws.cell(row=1, column=1).font
    data_font  = Font(name=hfont.name or "Calibri", size=hfont.size or 11)
    data_align = Alignment(vertical="center")

    model_to_first_sku: dict = {}
    for _, r in results_df.iterrows():
        mc  = str(r.get("model_code", "")).strip()
        sku = str(r.get("sku_num_sku_r3", "")).strip()
        if mc and sku and mc not in model_to_first_sku:
            model_to_first_sku[mc] = sku

    if df_cat is not None:
        exp_to_fullpath: dict = {}
        for _, _cr in df_cat.iterrows():
            _exp = str(_cr.get("export_category", "")).strip()
            _fp  = str(_cr.get("Category Path", "")).strip()
            if _exp and _fp and _exp not in exp_to_fullpath:
                exp_to_fullpath[_exp] = _fp
    else:
        exp_to_fullpath = {}

    if not ai_categories and df_cat is not None:
        _kw_cats = keyword_match_batch(results_df, df_cat)
    else:
        _kw_cats = None

    _brand_cache: dict = {}

    for i, (idx, src_row) in enumerate(results_df.iterrows()):
        row_idx  = i + 2
        row_data = {}

        for master_col, tmpl_col in MASTER_TO_TEMPLATE.items():
            val = src_row.get(master_col, "")
            if pd.notna(val) and str(val).strip() not in ("", "nan"):
                row_data[tmpl_col] = str(val).strip()

        # Long description with fallback
        row_data["Description"] = build_long_description(src_row)

        # Image validation and packing
        img_urls = []
        seen     = set()

        for c in IMAGE_COLS:
            if c in src_row.index:
                val = str(src_row[c]).strip()
                if val and val.lower() != "nan" and val.startswith(("http://", "https://")):
                    if val not in seen:
                        img_urls.append(val)
                        seen.add(val)

        if img_urls:
            main_candidate = img_urls[0]
            dim_status     = get_image_dimension_status(main_candidate)

            if dim_status["ok"]:
                row_data["MainImage"] = main_candidate
                for slot, url in zip(TEMPLATE_IMAGE_SLOTS[1:], img_urls[1:]):
                    row_data[slot] = url
            else:
                row_data["MainImage"] = FALLBACK_MAIN_IMAGE_URL
                # Do not export other images for this product

        mc = str(src_row.get("model_code", "")).strip()
        if mc and mc in model_to_first_sku:
            row_data["ParentSKU"] = model_to_first_sku[mc]

        gtin = _format_gtin(src_row.get("bar_code", ""))
        if gtin:
            row_data["GTIN_Barcode"] = gtin

        product_name = str(src_row.get("product_name", "")).strip()
        color_raw    = str(src_row.get("color", "")).strip()
        color        = color_raw.split("|")[0].strip()

        if color:
            row_data["Color"] = color.title()

        if product_name and color:
            if color.lower() not in product_name.lower():
                row_data["Name"] = f"{product_name} - {color.title()}"
            else:
                row_data["Name"] = product_name
        elif product_name:
            row_data["Name"] = product_name

        bw = str(src_row.get("business_weight", "")).strip()
        if bw and bw.lower() not in ("", "nan"):
            row_data["product_weight"] = re.sub(r'\s*kg\s*$', '', bw, flags=re.IGNORECASE).strip()

        size_val = re.sub(r'"+', '', str(src_row.get("size", ""))).strip().rstrip(".")
        if size_val.lower() not in ("", "nan", "no size"):
            pkg_name = row_data.get("Name", product_name)
            row_data["package_content"] = f"{pkg_name} - {size_val}"

        raw_brand = src_row.get("brand_name", "")
        if pd.notna(raw_brand) and str(raw_brand).strip():
            brand_key = str(raw_brand).strip()
            if brand_key not in _brand_cache:
                _brand_cache[brand_key] = match_brand(brand_key, df_brands)
            row_data["Brand"] = _brand_cache[brand_key]

        if ai_categories and i < len(ai_categories):
            primary_code, secondary_code = ai_categories[i]
        elif _kw_cats:
            primary_code, secondary_code = _kw_cats[i]
        else:
            primary_code, secondary_code = ("", "")

        primary_full   = exp_to_fullpath.get(primary_code, "")
        secondary_full = exp_to_fullpath.get(secondary_code, "")

        primary_code_only   = extract_export_code(primary_code)
        secondary_code_only = extract_export_code(secondary_code)

        if primary_code_only and primary_full:
            row_data["PrimaryCategory"] = f"{primary_code_only} - {primary_full}"
        elif primary_full:
            row_data["PrimaryCategory"] = primary_full
        elif primary_code:
            row_data["PrimaryCategory"] = primary_code

        if secondary_code_only and secondary_full:
            row_data["AdditionalCategory"] = f"{secondary_code_only} - {secondary_full}"
        elif secondary_full:
            row_data["AdditionalCategory"] = secondary_full
        elif secondary_code:
            row_data["AdditionalCategory"] = secondary_code

        per_row_override = (size_overrides or {}).get(idx)
        computed_var = get_variation(
            src_row,
            is_fashion=is_fashion,
            valid_sizes=valid_sizes,
            size_override=per_row_override,
        )

        if is_fashion:
            row_data["Size"] = computed_var
        else:
            row_data["Variation"] = computed_var

        row_data["Price_KES"] = "100000"
        row_data["Stock"]     = "0"

        if short_descs and i < len(short_descs) and short_descs[i]:
            row_data["short_description"] = short_descs[i]

        flag_red = is_fashion and is_size_missing(computed_var, valid_sizes or [])

        for tmpl_col, value in row_data.items():
            norm_tmpl_col = re.sub(r'[^a-z0-9]', '', str(tmpl_col).lower())

            if norm_tmpl_col not in header_map:
                current_max_col += 1
                header_cell       = ws.cell(row=1, column=current_max_col)
                header_cell.value = tmpl_col
                header_cell.font  = Font(bold=True)
                header_map[norm_tmpl_col] = current_max_col

            cell           = ws.cell(row=row_idx, column=header_map[norm_tmpl_col])
            cell.value     = value
            cell.font      = data_font
            cell.alignment = data_align

            if flag_red and tmpl_col in ("Size", "Variation"):
                cell.fill = RED_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.header("Master Data")
    _bundled_exists = any(
        os.path.exists(p) for p in [MASTER_PATH, MASTER_PATH.replace(".xlsx", ".csv")]
    )
    if _bundled_exists:
        st.success("Master file loaded from system")
        with st.expander("Override with a different file"):
            uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx", "csv"])
    else:
        uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx", "csv"])

    st.markdown("---")
    _WORKING_STATE_KEYS = {
        "run_id", "size_overrides", "cat_overrides",
        "ai_categories", "short_descs", "combined",
    }

    if st.button(
        "Clear Working Data",
        use_container_width=True,
        help="Resets search results and overrides. Category and brand files stay loaded."
    ):
        for key in list(st.session_state.keys()):
            if key in _WORKING_STATE_KEYS or key.startswith(("size_fix_", "prim_", "cat_search")):
                del st.session_state[key]
        st.rerun()

    st.markdown("---")
    st.header("Category Matching")
    use_ai_matching = st.toggle(
        "AI matching",
        value=False,
        help="OFF = fast keyword/TF-IDF. ON = TF-IDF shortlist + AI gateway rerank.",
    )

    if use_ai_matching:
        st.markdown('<span class="ai-badge">AI MODE ON</span>', unsafe_allow_html=True)
        show_key = st.checkbox("Show key while typing", value=False)
        ai_gateway_api_key = st.text_input(
            "AI Gateway API key",
            type="default" if show_key else "password",
            value=os.environ.get("AI_GATEWAY_API_KEY", HARDCODED_AI_GATEWAY_KEY),
            placeholder="Paste your gateway key here",
        )

        ai_model = st.selectbox(
            "Model",
            ["gpt-4o-mini", "gpt-5", "gemini-2.5-pro", "llama-3.1-8b-instant",
             "llama-3.3-70b-versatile", "mixtral-8x7b-32768"],
            index=0,
        )

        shortlist_k  = st.slider("Shortlist size", 10, 50, 30)
        concurrency  = st.slider("Parallel AI requests", 1, 30, 10)
        st.markdown("---")
        ai_short_desc = st.toggle("AI short descriptions", value=True)
        debug_ai      = st.checkbox(
            "Debug AI gateway",
            value=DEBUG_AI_DEFAULT,
            help="Shows raw gateway errors, status codes, and response snippets.",
        )
        st.caption(f"Gateway URL: {AI_GATEWAY_BASE_URL}")
    else:
        st.markdown('<span class="kw-badge">KEYWORD MODE</span>', unsafe_allow_html=True)
        st.caption("Instant vectorised TF-IDF keyword matching. No API key needed.")
        ai_gateway_api_key = ""
        ai_model           = DEFAULT_AI_MODEL
        shortlist_k        = 30
        concurrency        = 10
        ai_short_desc      = False
        debug_ai           = False

    st.markdown("---")
    st.header("Product Type")
    product_type = st.radio(
        "Product type",
        ["Fashion", "Other"],
        index=0,
        horizontal=True,
        help=(
            "Fashion: uses the size column with UK-size extraction + sizes.txt validation.\n"
            "Other: uses the size column directly; ... only when size is empty."
        ),
    )
    is_fashion = product_type == "Fashion"

    valid_sizes: list = parse_valid_sizes(SIZES_PATH)
    if valid_sizes:
        st.sidebar.info(f"sizes.txt loaded: {len(valid_sizes)} sizes")
    else:
        st.sidebar.warning("sizes.txt not found in project folder.")

    st.markdown("---")
    st.header("Search Fields")
    also_search_name = st.checkbox("Also search by product name", value=False)

# =============================================================================
# LOAD REFERENCE DATA
# =============================================================================

try:
    ref_bytes = open(DECA_CAT_PATH, "rb").read()
    st.sidebar.success("deca_cat.xlsx loaded")
except FileNotFoundError:
    ref_bytes = None
    st.sidebar.error(f"`{DECA_CAT_PATH}` not found. Place it alongside app.py and restart.")

if ref_bytes:
    df_cat, df_brands = load_reference_data(ref_bytes)
    st.sidebar.success(f"{len(df_cat):,} categories · {len(df_brands)} brands")
    leaves, vectorizer, tfidf_matrix, path_to_export = build_tfidf_index(ref_bytes)
else:
    df_cat = df_brands = leaves = vectorizer = tfidf_matrix = path_to_export = None

# =============================================================================
# LOAD MASTER DATA
# =============================================================================

if uploaded_master:
    master_bytes = uploaded_master.read()
    is_csv       = uploaded_master.name.endswith(".csv")
    df_master    = load_master(master_bytes, is_csv)
    st.sidebar.success(f"{len(df_master):,} product rows loaded")
else:
    df_master = load_master_fast()
    if df_master is None:
        st.error(f"Master file not found. Place '{MASTER_PATH}' in the same folder as app.py.")
        st.stop()
    st.sidebar.info(f"Bundled master · {len(df_master):,} rows")

img_cols_present = [c for c in IMAGE_COLS if c in df_master.columns]
data_cols        = [c for c in df_master.columns if c not in img_cols_present]

# =============================================================================
# SEARCH
# =============================================================================

def search(q: str) -> pd.DataFrame:
    mask = pd.Series(False, index=df_master.index)
    mask |= df_master["sku_num_sku_r3"].fillna("").str.strip() == q.strip()
    if also_search_name and "product_name" in df_master.columns:
        mask |= df_master["product_name"].fillna("").str.lower().str.contains(q.lower(), regex=False)
    return df_master[mask].copy()

# =============================================================================
# INPUT TABS
# =============================================================================

tab1, tab2, tab3 = st.tabs(["Upload a List", "Manual Entry", "Explore Categories"])
queries = []

with tab1:
    uploaded_list = st.file_uploader(
        "Upload file with SKU numbers",
        type=["xlsx", "csv", "txt"],
        help="One SKU per row. For Excel or CSV, SKUs must be in column A.",
    )
    if uploaded_list:
        ext = uploaded_list.name.rsplit(".", 1)[-1].lower()
        if ext == "txt":
            queries = [l.strip() for l in uploaded_list.read().decode().splitlines() if l.strip()]
        elif ext == "csv":
            q_df    = pd.read_csv(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        else:
            q_df    = pd.read_excel(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        st.success(f"Loaded **{len(queries)}** search terms")

with tab2:
    manual = st.text_area(
        "Enter one SKU number per line",
        height=160,
        placeholder="4273417\n4273418\n4273423",
    )
    manual_submitted = st.button("Search SKUs", type="primary", use_container_width=True)
    if manual_submitted and manual.strip():
        st.session_state["manual_queries"] = [q.strip() for q in manual.strip().splitlines() if q.strip()]
    if "manual_queries" in st.session_state and not manual_submitted:
        queries = st.session_state["manual_queries"]
    elif manual_submitted and manual.strip():
        queries = st.session_state["manual_queries"]

with tab3:
    st.subheader("Category Explorer")
    if df_cat is None:
        st.warning("deca_cat.xlsx not loaded, categories unavailable.")
    else:
        cat_display = df_cat[["Category Path", "export_category", "category_name"]].copy()
        cat_display.columns = ["Full Path", "Export Code", "Category Name"]
        cat_display = cat_display.drop_duplicates(subset=["Export Code"]).reset_index(drop=True)

        path_parts = cat_display["Full Path"].str.split("/", expand=True)
        n_levels   = path_parts.shape[1]
        path_parts.columns = [f"L{i+1}" for i in range(n_levels)]
        cat_display = pd.concat([cat_display, path_parts], axis=1)

        total_cats = len(cat_display)
        n_l1       = cat_display["L1"].nunique()
        deepest    = cat_display["Full Path"].str.count("/").max() + 1
        sc1, sc2, sc3 = st.columns(3)
        sc1.metric("Total categories", total_cats)
        sc2.metric("Top-level groups", n_l1)
        sc3.metric("Max depth", deepest)

        st.markdown("---")

        view_mode = st.radio(
            "View as",
            ["Tree (drill-down)", "Flat table"],
            horizontal=True,
            key="cat_view_mode",
        )

        cat_explore_search = st.text_input(
            "Search categories",
            placeholder="e.g. running, football, kids, hiking",
            key="cat_explore_search",
        )
        q_lower = cat_explore_search.strip().lower()

        if q_lower:
            mask = (
                cat_display["Full Path"].str.lower().str.contains(q_lower, na=False)
                | cat_display["Category Name"].str.lower().str.contains(q_lower, na=False)
                | cat_display["Export Code"].str.lower().str.contains(q_lower, na=False)
            )
            filtered = cat_display[mask].reset_index(drop=True)
        else:
            filtered = cat_display.copy()

        st.caption(
            f"Showing **{len(filtered)}** / {total_cats} categor{'y' if len(filtered)==1 else 'ies'}"
            + (f" matching '{cat_explore_search}'" if q_lower else "")
        )

        if view_mode == "Tree (drill-down)":
            l1_options = sorted(filtered["L1"].dropna().unique().tolist())
            if not l1_options:
                st.info("No categories match your search.")
            else:
                selected_l1 = st.selectbox(
                    "Top-level group (L1)",
                    ["(all)"] + l1_options,
                    key="cat_tree_l1",
                )
                sub = filtered if selected_l1 == "(all)" else filtered[filtered["L1"] == selected_l1]

                if selected_l1 != "(all)" and "L2" in sub.columns:
                    l2_options = sorted(sub["L2"].dropna().unique().tolist())
                    selected_l2 = st.selectbox(
                        "Sub-group (L2)",
                        ["(all)"] + l2_options,
                        key="cat_tree_l2",
                    )
                    if selected_l2 != "(all)":
                        sub = sub[sub["L2"] == selected_l2]

                        if "L3" in sub.columns:
                            l3_options = sorted(sub["L3"].dropna().unique().tolist())
                            if l3_options:
                                selected_l3 = st.selectbox(
                                    "Category (L3)",
                                    ["(all)"] + l3_options,
                                    key="cat_tree_l3",
                                )
                                if selected_l3 != "(all)":
                                    sub = sub[sub["L3"] == selected_l3]

                st.caption(f"{len(sub)} categor{'y' if len(sub)==1 else 'ies'} in selection")

                level_cols = [c for c in [f"L{i+1}" for i in range(n_levels)] if c in sub.columns]
                for l1_val, grp_l1 in sub.groupby("L1", sort=True):
                    with st.expander(f"{l1_val}  ({len(grp_l1)})", expanded=(len(l1_options) == 1 or bool(q_lower))):
                        if "L2" in grp_l1.columns:
                            for l2_val, grp_l2 in grp_l1.groupby("L2", sort=True, dropna=False):
                                l2_label = str(l2_val) if pd.notna(l2_val) else "(no sub-group)"
                                st.markdown(f"**{l2_label}** - {len(grp_l2)} item(s)")
                                rows_md = []
                                for _, r in grp_l2.iterrows():
                                    deeper = " / ".join(
                                        str(r[c]) for c in level_cols[2:]
                                        if pd.notna(r.get(c)) and str(r.get(c)).strip()
                                    )
                                    label = deeper if deeper else r["Category Name"]
                                    rows_md.append(f"- `{r['Export Code']}` &nbsp; {label}")
                                st.markdown("\n".join(rows_md), unsafe_allow_html=True)
                        else:
                            rows_md = [
                                f"- `{r['Export Code']}` &nbsp; {r['Category Name']}"
                                for _, r in grp_l1.iterrows()
                            ]
                            st.markdown("\n".join(rows_md), unsafe_allow_html=True)
        else:
            level_cols = [c for c in [f"L{i+1}" for i in range(n_levels)] if c in filtered.columns]
            display_df = pd.concat([filtered[level_cols], filtered[["Export Code"]]], axis=1)
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                height=480,
                column_config={
                    col: st.column_config.TextColumn(col, width="medium")
                    for col in display_df.columns
                },
            )

        st.markdown("---")
        cat_out = io.BytesIO()
        with pd.ExcelWriter(cat_out, engine="openpyxl") as w:
            cat_display[["Full Path", "Export Code", "Category Name"]].to_excel(
                w, index=False, sheet_name="Categories"
            )
        st.download_button(
            "Download full category list (.xlsx)",
            data=cat_out.getvalue(),
            file_name="decathlon_categories.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# =============================================================================
# RESULTS
# =============================================================================

if queries:
    st.markdown("---")
    all_result_frames = []
    no_match          = []

    for q in queries:
        res = search(q)
        if res.empty:
            no_match.append(q)
        else:
            res.insert(0, "Search Term", q)
            all_result_frames.append((q, res))

    if no_match:
        st.warning(f"No matches found for: **{', '.join(no_match)}**")

    if all_result_frames:
        total_rows = sum(len(r) for _, r in all_result_frames)
        st.success(f"**{total_rows} rows** matched across **{len(all_result_frames)}** query(ies)")

        combined = pd.concat([r for _, r in all_result_frames], ignore_index=True)

        ai_categories = None
        ai_cat_debug  = []

        if df_cat is not None and use_ai_matching and ai_gateway_api_key:
            n               = len(combined)
            unique_models_n = combined["model_code"].nunique() if "model_code" in combined.columns else n
            est             = max(2, unique_models_n // concurrency + 2)
            with st.spinner(f"AI category matching {unique_models_n} unique models (~{est}s)..."):
                try:
                    ai_categories, _model_cats, ai_cat_debug = ai_match_categories(
                        combined, leaves, vectorizer, tfidf_matrix, path_to_export,
                        ai_gateway_api_key, ai_model, shortlist_k, concurrency, debug_ai,
                    )
                    st.success(f"AI matched {unique_models_n} models to {n} SKUs")
                except Exception as e:
                    st.error(f"AI category error: {e}")
                    use_ai_matching = False
        elif df_cat is not None and use_ai_matching and not ai_gateway_api_key:
            st.warning("Enter your AI Gateway API key in the sidebar to use AI matching.")
            use_ai_matching = False

        short_descs  = None
        ai_desc_debug = []

        if use_ai_matching and ai_short_desc and ai_gateway_api_key:
            with st.spinner(f"Generating AI short descriptions ({len(combined)} products)..."):
                try:
                    short_descs, ai_desc_debug = ai_short_descriptions(
                        combined, ai_gateway_api_key, ai_model, concurrency, debug_ai
                    )
                    st.success("Short descriptions generated")
                except Exception as e:
                    st.error(f"Short desc error: {e}")
                    short_descs   = None
                    ai_desc_debug = []

        if debug_ai and (ai_cat_debug or ai_desc_debug):
            st.markdown("---")
            st.subheader("AI Gateway Debug")

            debug_rows = ai_cat_debug + ai_desc_debug

            summary_df = pd.DataFrame([
                {
                    "model_code":   r.get("model_code", ""),
                    "task_type":    r.get("task_type", ""),
                    "error":        r.get("error", ""),
                    "status":       (r.get("debug") or {}).get("status"),
                    "error_type":   (r.get("debug") or {}).get("error_type"),
                    "used_fallback":(r.get("debug") or {}).get("used_fallback"),
                    "model":        (r.get("debug") or {}).get("model"),
                }
                for r in debug_rows
            ])

            if not summary_df.empty:
                st.dataframe(summary_df, use_container_width=True, hide_index=True)

            for i, row in enumerate(debug_rows, 1):
                dbg = row.get("debug", {}) or {}
                with st.expander(
                    f"Debug #{i} | {row.get('task_type','')} | model_code={row.get('model_code','')} | {row.get('error','')[:80]}"
                ):
                    st.markdown(f"**Error:** `{row.get('error', '')}`")
                    st.markdown(f"**Query preview:** `{row.get('query_preview', '')}`")
                    st.markdown(f"**Candidate count:** `{row.get('candidate_count', 0)}`")
                    st.markdown(f"**URL:** `{dbg.get('url', '')}`")
                    st.markdown(f"**Status:** `{dbg.get('status', '')}`")
                    st.markdown(f"**Error type:** `{dbg.get('error_type', '')}`")
                    st.markdown(f"**Error message:** `{dbg.get('error_message', '')}`")
                    st.markdown(f"**Used fallback:** `{dbg.get('used_fallback', False)}`")
                    st.markdown(f"**Model:** `{dbg.get('model', '')}`")

                    if dbg.get("candidate_preview"):
                        st.markdown("**Candidate preview:**")
                        st.code(json.dumps(dbg.get("candidate_preview", []), indent=2), language="json")

                    if dbg.get("request_payload") is not None:
                        st.markdown("**Request payload:**")
                        st.code(json.dumps(dbg.get("request_payload"), indent=2), language="json")

                    if dbg.get("raw_model_content"):
                        st.markdown("**Raw model content:**")
                        st.code(dbg.get("raw_model_content", ""), language="text")

                    if dbg.get("body_text"):
                        st.markdown("**Gateway response body:**")
                        st.code(dbg.get("body_text", ""), language="text")

        if short_descs is None:
            short_descs = [rule_based_short_desc(row) for _, row in combined.iterrows()]

        if df_cat is not None:
            _exp_to_path: dict = {}
            for _, _rc in df_cat.iterrows():
                _e = str(_rc.get("export_category", "")).strip()
                _p = str(_rc.get("Category Path", "")).strip()
                if _e and _p and _e not in _exp_to_path:
                    _exp_to_path[_e] = _p
        else:
            _exp_to_path = {}

        def _code_to_full(code):
            c = str(code).strip()
            if not c:
                return ""
            p         = _exp_to_path.get(c, "")
            code_only = extract_export_code(c)
            if code_only and p:
                return f"{code_only} - {p}"
            return p or c

        if "size_overrides" not in st.session_state:
            st.session_state.size_overrides = {}

        preview = combined.copy()
        preview.reset_index(drop=True, inplace=True)

        if ai_categories:
            preview["_primary_cat"] = [_code_to_full(c[0]) for c in ai_categories]
        elif df_cat is not None:
            kw = keyword_match_batch(preview, df_cat)
            preview["_primary_cat"] = [_code_to_full(c[0]) for c in kw]
        else:
            preview["_primary_cat"] = ""

        def _compute_var(row):
            override = st.session_state.size_overrides.get(row.name)
            return get_variation(
                row,
                is_fashion=is_fashion,
                valid_sizes=valid_sizes,
                size_override=override,
            )

        preview["_variation"] = preview.apply(_compute_var, axis=1)
        preview["_size_ok"]   = preview["_variation"].apply(
            lambda v: not is_size_missing(v, valid_sizes) if is_fashion else True
        )

        st.markdown("---")
        st.subheader(f"Export Preview - {total_rows} SKU(s)")

        if is_fashion:
            st.info(
                "**Fashion mode** - rows highlighted in red have a size not found in sizes.txt. "
                "Use the dropdowns below to fix them before downloading."
            )
        else:
            st.info("**Other mode** - variation is taken from the size column; '...' shown when empty.")

        def _export_name(row):
            pn  = str(row.get("product_name", "")).strip()
            col = str(row.get("color", "")).split("|")[0].strip()
            if pn and col and col.lower() not in pn.lower():
                return f"{pn} - {col.title()}"
            return pn

        preview["_export_name"] = preview.apply(_export_name, axis=1)

        if is_fashion:
            preview["Size"]   = preview["_variation"]
            display_cols = ["sku_num_sku_r3", "_export_name", "color", "Size",
                            "_primary_cat", "brand_name", "bar_code", "_size_ok"]
        else:
            preview["Variation"] = preview["_variation"]
            display_cols = ["sku_num_sku_r3", "_export_name", "color", "Variation",
                            "_primary_cat", "brand_name", "bar_code", "_size_ok"]

        show_cols = [c for c in display_cols if c in preview.columns]

        col_cfg = {
            "sku_num_sku_r3": st.column_config.TextColumn("SKU", width="small"),
            "_export_name":   st.column_config.TextColumn("Name (export)", width="large"),
            "color":          st.column_config.TextColumn("Colour", width="medium"),
            "Size":           st.column_config.TextColumn("Size (Export)", width="medium"),
            "Variation":      st.column_config.TextColumn("Variation (Export)", width="medium"),
            "_primary_cat":   st.column_config.TextColumn("Primary Category", width="large"),
            "brand_name":     st.column_config.TextColumn("Brand", width="small"),
            "bar_code":       st.column_config.TextColumn("Barcode", width="medium"),
            "_size_ok":       st.column_config.CheckboxColumn("Size OK", width="small"),
        }

        df_display = preview[show_cols].copy()
        if is_fashion and "_size_ok" in df_display.columns:
            df_display.insert(0, "⚠️", df_display["_size_ok"].apply(
                lambda ok: "" if ok else "⚠️ fix size"
            ))
            col_cfg["⚠️"] = st.column_config.TextColumn("", width="small")

        st.dataframe(df_display, use_container_width=True, hide_index=True, height=420, column_config=col_cfg)

        if is_fashion and valid_sizes:
            bad_rows = preview[~preview["_size_ok"]]
            if not bad_rows.empty:
                st.markdown("---")
                st.subheader(f"Fix Sizes - {len(bad_rows)} row(s) need attention")
                st.caption("Select the correct size for each flagged row. Changes apply to the downloaded template.")

                for pos_idx, row in bad_rows.iterrows():
                    sku     = row.get("sku_num_sku_r3", f"row {pos_idx}")
                    name    = str(row.get("product_name", ""))[:50]
                    raw_s   = str(row.get("size", ""))
                    current_override = st.session_state.size_overrides.get(pos_idx)
                    current_val      = current_override or row["_variation"]

                    opts = ["(auto)"] + valid_sizes
                    try:
                        sel_idx = opts.index(current_val) if current_val in opts else 0
                    except ValueError:
                        sel_idx = 0

                    col1, col2, col3 = st.columns([2, 2, 3])
                    col1.markdown(f"**{sku}** \n`{name}`")
                    col2.markdown(f"Master: `{raw_s}`  \nAuto: `{row['_variation']}`")
                    chosen = col3.selectbox(
                        f"Size for {sku}",
                        opts,
                        index=sel_idx,
                        key=f"size_fix_{pos_idx}",
                        label_visibility="collapsed",
                    )
                    if chosen != "(auto)":
                        st.session_state.size_overrides[pos_idx] = chosen
                    elif pos_idx in st.session_state.size_overrides:
                        del st.session_state.size_overrides[pos_idx]

                if st.button("Re-apply size fixes to preview"):
                    st.rerun()
            else:
                st.success("All sizes matched in sizes.txt")

        if df_cat is not None:
            st.markdown("---")
            mode_label = "AI" if (use_ai_matching and ai_categories) else "Keyword"
            st.subheader(f"Category Editor - {mode_label}")
            st.caption(
                "Categories are shared across all SKUs with the same model code. "
                "Edit one row per model, siblings update automatically on export."
            )

            export_to_path: dict = {}
            for _, row_c in df_cat.iterrows():
                exp  = str(row_c.get("export_category", "")).strip()
                path = str(row_c.get("Category Path", "")).strip()
                if exp and path and exp not in export_to_path:
                    export_to_path[exp] = path
            path_label_to_export: dict = {v: k for k, v in export_to_path.items()}

            def export_to_label(code: str) -> str:
                if not code:
                    return ""
                path      = export_to_path.get(code, "")
                code_only = extract_export_code(code)
                if code_only and path:
                    return f"{code_only} - {path}"
                return path or code

            def label_to_export(label: str) -> str:
                if not label or label == "(auto)":
                    return ""
                return path_label_to_export.get(label, label)

            all_path_labels    = sorted(export_to_path.values())
            all_labels_w_blank = ["(auto)"] + all_path_labels

            if "cat_overrides" not in st.session_state:
                st.session_state.cat_overrides = {}

            cat_search = st.text_input(
                "Filter category list",
                placeholder="e.g. football, running, kids...",
                key="cat_search",
            )
            q_cat = cat_search.strip().lower()
            filtered_labels = (
                ["(auto)"] + [lbl for lbl in all_path_labels if q_cat in lbl.lower()]
                if q_cat else all_labels_w_blank
            )
            st.caption(
                f"{len(filtered_labels)-1} categories shown"
                + (f" matching '{cat_search}'" if q_cat else " (all)")
                + f" · {len(st.session_state.cat_overrides)} model override(s)"
            )
            st.markdown("---")

            seen_models: set = set()
            hc1, hc2, hc4   = st.columns([2, 5, 1])
            hc1.markdown("**Model · SKUs**")
            hc2.markdown("**Primary Category**")
            hc4.markdown("**Method**")

            for i, (_, prow) in enumerate(combined.iterrows()):
                mc = str(prow.get("model_code", "")).strip()
                if mc in seen_models:
                    continue
                seen_models.add(mc)

                sku_count = len(combined[combined["model_code"] == mc])
                name      = str(prow.get("product_name", ""))[:40]

                if ai_categories:
                    first_idx = next(
                        j for j, (_, r) in enumerate(combined.iterrows())
                        if str(r.get("model_code", "")).strip() == mc
                    )
                    auto_prim_code, _ = ai_categories[first_idx]
                else:
                    auto_prim_code, _ = keyword_match_category(prow, df_cat)

                auto_prim_label = export_to_label(auto_prim_code)
                override        = st.session_state.cat_overrides.get(mc, {})
                cur_prim_label  = export_to_label(override.get("primary", auto_prim_code)) or auto_prim_label

                c1, c2, c4 = st.columns([2, 5, 1])
                c1.markdown(f"**{mc}** \n{name} \n`{sku_count} SKU(s)`")

                prim_opts = (
                    filtered_labels if cur_prim_label in filtered_labels
                    else ["(auto)", cur_prim_label] + [l for l in filtered_labels if l != "(auto)"]
                )
                try:
                    prim_idx = prim_opts.index(cur_prim_label)
                except ValueError:
                    prim_idx = 0

                new_prim_label = c2.selectbox(
                    f"Primary #{mc}",
                    prim_opts,
                    index=prim_idx,
                    label_visibility="collapsed",
                    key=f"prim_{mc}",
                )
                new_prim_code = label_to_export(new_prim_label) if new_prim_label != "(auto)" else auto_prim_code

                if new_prim_label != "(auto)":
                    st.session_state.cat_overrides[mc] = {"primary": new_prim_code, "additional": ""}
                elif mc in st.session_state.cat_overrides:
                    del st.session_state.cat_overrides[mc]

                badge = "Manual" if mc in st.session_state.cat_overrides else (
                    "AI" if (use_ai_matching and ai_categories) else "Keyword"
                )
                c4.markdown(f"`{badge}`")

        st.markdown("---")

        if df_cat is None:
            st.warning("deca_cat.xlsx not loaded, template download unavailable.")
        else:
            try:
                merged_cats = []
                for i2, (_, prow) in enumerate(combined.iterrows()):
                    mc2      = str(prow.get("model_code", "")).strip()
                    override = st.session_state.get("cat_overrides", {}).get(mc2)
                    if override:
                        merged_cats.append((override["primary"], override.get("additional", "")))
                    elif ai_categories:
                        first_idx = next(
                            j for j, (_, r) in enumerate(combined.iterrows())
                            if str(r.get("model_code", "")).strip() == mc2
                        )
                        merged_cats.append(ai_categories[first_idx])
                    else:
                        merged_cats.append(keyword_match_category(prow, df_cat))

                idx_overrides = {}
                for pos_idx, sz in st.session_state.size_overrides.items():
                    if pos_idx < len(combined):
                        real_idx = combined.index[pos_idx]
                        idx_overrides[real_idx] = sz

                tpl_bytes = build_template(
                    combined,
                    df_cat,
                    df_brands,
                    ai_categories=merged_cats,
                    short_descs=short_descs,
                    is_fashion=is_fashion,
                    valid_sizes=valid_sizes,
                    size_overrides=idx_overrides,
                )
                mode_icon = "AI" if (use_ai_matching and ai_categories) else "KW"
                st.download_button(
                    f"{mode_icon} Download Filled Upload Template (.xlsx)",
                    data=tpl_bytes,
                    file_name="decathlon_upload_template_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            except FileNotFoundError:
                st.warning(
                    "Template file not found. "
                    "Place `product-creation-template.xlsx` in the app folder."
                )

else:
    st.info("Upload a list or type search terms above to get started.")

st.markdown("---")
st.caption("Decathlon Product Lookup · Powered by your Decathlon working file")
